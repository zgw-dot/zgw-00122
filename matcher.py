import csv
import io
import json
from datetime import datetime, timezone, timedelta
import hashlib
from models import (
    db, Batch, PurchaseOrder, Invoice, MatchResult, ExceptionItem,
    ToleranceHistory, AuditLog, PayableRecalcNote, NoteComparison,
    ImportDraft, ImportDraftIssue, ImportPlan, PlanSnapshot,
    HealthCheckRule, HealthCheckHistory, HealthCheckResult,
    HandoverList, HandoverListItem,
    ReleasePackage, ReleasePackageItem,
    BATCH_STATUS_VALIDATING, BATCH_STATUS_MATCHED, BATCH_STATUS_EXCEPTION,
    BATCH_STATUS_FAILED, BATCH_STATUS_CREATED, BATCH_STATUS_CONFIRMED,
    MATCH_TYPE_EXACT, MATCH_TYPE_TOLERANCE, MATCH_TYPE_OVER_TOLERANCE,
    MATCH_TYPE_UNMATCHED_PO, MATCH_TYPE_UNMATCHED_INVOICE,
    EXCEPTION_MISSING_FIELD, EXCEPTION_OVER_TOLERANCE,
    EXCEPTION_DUPLICATE_INVOICE,
    EXCEPTION_STATUS_PENDING, RESULT_STATUS_PENDING, RESULT_STATUS_REJECTED,
    REVIEW_STATUS_PENDING, REVIEW_STATUS_CONFIRMED, REVIEW_STATUS_IGNORED,
    DRAFT_STATUS_PENDING, DRAFT_STATUS_CONFIRMED, DRAFT_STATUS_DISCARDED,
    DRAFT_STATUS_CONFLICT, DRAFT_STATUS_CANCELLED,
    DRAFT_FILE_TYPE_PO, DRAFT_FILE_TYPE_INVOICE,
    PLAN_STATUS_PENDING, PLAN_STATUS_CONFIRMED, PLAN_STATUS_CANCELLED, PLAN_STATUS_UNDONE,
    DRAFT_EXPIRE_HOURS,
    PRECHECK_ERROR, PRECHECK_WARNING, PRECHECK_INFO,
    ROW_ACTION_ADD, ROW_ACTION_OVERWRITE, ROW_ACTION_SKIP, ROW_ACTION_CONFLICT,
    HEALTH_SEVERITY_BLOCKER, HEALTH_SEVERITY_WARNING, HEALTH_SEVERITY_INFO,
    HEALTH_RULE_DUPLICATE_PO, HEALTH_RULE_DUPLICATE_INVOICE,
    HEALTH_RULE_MISSING_COLUMNS, HEALTH_RULE_NEGATIVE_AMOUNT,
    HEALTH_RULE_VENDOR_MISMATCH, HEALTH_RULE_CONFIRMED_OVERRIDE_RISK,
    DEFAULT_HEALTH_RULES,
    HANDOVER_STATUS_DRAFT, HANDOVER_STATUS_COMPLETED, HANDOVER_STATUS_VOID,
    HANDOVER_PERMISSION_COMPLETE, HANDOVER_PERMISSION_VOID,
    HANDOVER_ROLE_ADMIN, HANDOVER_ROLE_FINANCE_LEAD,
    compute_rule_version, compute_note_content_hash, compute_health_rule_version,
    compute_handover_content_hash, compute_release_content_hash,
    RELEASE_STATUS_DRAFT, RELEASE_STATUS_PENDING, RELEASE_STATUS_APPROVED,
    RELEASE_STATUS_REJECTED, RELEASE_STATUS_REVOKED, RELEASE_STATUS_EXPIRED,
    RELEASE_PERMISSION_APPROVE, RELEASE_PERMISSION_REJECT, RELEASE_PERMISSION_REVOKE,
    RELEASE_PERMISSION_CREATE, RELEASE_PERMISSION_VIEW,
    HANDOVER_ROLE_ADMIN, HANDOVER_ROLE_FINANCE_LEAD, HANDOVER_ROLE_FINANCE,
)

PO_REQUIRED_COLUMNS = ["po_number", "vendor_code", "vendor_name", "amount", "po_date"]
INVOICE_REQUIRED_COLUMNS = ["invoice_number", "vendor_code", "vendor_name", "amount", "invoice_date"]


class ValidationError(Exception):
    def __init__(self, errors):
        self.errors = errors
        self.details = errors if isinstance(errors, list) else [errors]
        super().__init__("; ".join(self.details))


def _extract_storage(file_storage, fallback_filename=None):
    """Return (read_text_fn, filename, is_excel) from either FileStorage or (text, filename) tuple."""
    if isinstance(file_storage, tuple):
        content, fname = file_storage
        fname = fallback_filename or fname or ""

        def read_text():
            return content

        return read_text, fname, fname.endswith((".xlsx", ".xls"))
    # werkzeug FileStorage or similar
    fname = getattr(file_storage, "filename", None) or fallback_filename or ""

    def read_text():
        raw = file_storage.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig")
        return raw

    return read_text, fname, fname.endswith((".xlsx", ".xls"))


def parse_csv_file(file_storage):
    read_text, _, _ = _extract_storage(file_storage)
    raw = read_text()
    reader = csv.DictReader(io.StringIO(raw))
    rows = []
    headers = reader.fieldnames or []
    for row in reader:
        rows.append(row)
    return headers, rows


def parse_excel_file(file_storage):
    from openpyxl import load_workbook
    read_text, _, _ = _extract_storage(file_storage)
    raw = read_text()
    if isinstance(raw, str):
        raw = raw.encode("utf-8")
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
    rows = []
    for row in rows_iter:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = str(val) if val is not None else ""
        rows.append(d)
    wb.close()
    return headers, rows


def parse_file(file_storage, filename=None):
    _, fname, is_excel = _extract_storage(file_storage, fallback_filename=filename)
    if is_excel:
        return parse_excel_file(file_storage)
    return parse_csv_file(file_storage)


def validate_columns(headers, required):
    lower_headers = [h.strip().lower() for h in headers]
    missing = [c for c in required if c.lower() not in lower_headers]
    return missing


def validate_po_row(row, row_num):
    errors = []
    for col in PO_REQUIRED_COLUMNS:
        val = row.get(col, "").strip() if row.get(col) else ""
        if not val:
            errors.append(f"第{row_num}行: 缺少必填字段 '{col}'")
    if row.get("amount"):
        try:
            float(row["amount"])
        except (ValueError, TypeError):
            errors.append(f"第{row_num}行: 金额格式错误 '{row['amount']}'")
    return errors


def validate_invoice_row(row, row_num):
    errors = []
    for col in INVOICE_REQUIRED_COLUMNS:
        val = row.get(col, "").strip() if row.get(col) else ""
        if not val:
            errors.append(f"第{row_num}行: 缺少必填字段 '{col}'")
    if row.get("amount"):
        try:
            float(row["amount"])
        except (ValueError, TypeError):
            errors.append(f"第{row_num}行: 金额格式错误 '{row['amount']}'")
    return errors


def check_duplicate_invoices(rows):
    seen = {}
    duplicates = []
    for i, row in enumerate(rows, 2):
        inv_num = row.get("invoice_number", "").strip()
        if inv_num in seen:
            duplicates.append(f"发票号重复: '{inv_num}' 出现在第{seen[inv_num]}行和第{i}行")
        else:
            seen[inv_num] = i
    return duplicates


def import_purchase_orders(batch, rows):
    for i, row in enumerate(rows, 2):
        po = PurchaseOrder(
            batch_id=batch.id,
            po_number=row.get("po_number", "").strip(),
            vendor_code=row.get("vendor_code", "").strip(),
            vendor_name=row.get("vendor_name", "").strip(),
            amount=float(row.get("amount", 0)),
            currency=row.get("currency", "CNY").strip(),
            po_date=row.get("po_date", "").strip(),
            raw_data=json.dumps(row, ensure_ascii=False),
        )
        db.session.add(po)


def import_invoices(batch, rows):
    for i, row in enumerate(rows, 2):
        inv = Invoice(
            batch_id=batch.id,
            invoice_number=row.get("invoice_number", "").strip(),
            vendor_code=row.get("vendor_code", "").strip(),
            vendor_name=row.get("vendor_name", "").strip(),
            amount=float(row.get("amount", 0)),
            currency=row.get("currency", "CNY").strip(),
            invoice_date=row.get("invoice_date", "").strip(),
            raw_data=json.dumps(row, ensure_ascii=False),
        )
        db.session.add(inv)


def perform_matching(batch):
    tolerance_pct = batch.tolerance_pct
    tolerance_abs = batch.tolerance_abs
    rule_version = compute_rule_version(tolerance_pct, tolerance_abs)
    batch.rule_version = rule_version

    pos = PurchaseOrder.query.filter_by(batch_id=batch.id).all()
    invoices = Invoice.query.filter_by(batch_id=batch.id).all()

    inv_by_vendor = {}
    for inv in invoices:
        inv_by_vendor.setdefault(inv.vendor_code, []).append(inv)

    matched_inv_ids = set()
    has_exceptions = False

    for po in pos:
        vendor_invs = inv_by_vendor.get(po.vendor_code, [])
        best_match = None
        best_diff = float("inf")
        best_match_type = None

        for inv in vendor_invs:
            if inv.id in matched_inv_ids:
                continue
            diff = abs(po.amount - inv.amount)
            if po.amount == inv.amount:
                if best_match_type != MATCH_TYPE_EXACT:
                    best_match = inv
                    best_diff = diff
                    best_match_type = MATCH_TYPE_EXACT
            elif diff <= tolerance_abs or (po.amount > 0 and diff / po.amount * 100 <= tolerance_pct):
                if best_match_type != MATCH_TYPE_EXACT and diff < best_diff:
                    best_match = inv
                    best_diff = diff
                    best_match_type = MATCH_TYPE_TOLERANCE

        if best_match:
            matched_inv_ids.add(best_match.id)
            is_exc = best_match_type == MATCH_TYPE_TOLERANCE
            mr = MatchResult(
                batch_id=batch.id,
                po_id=po.id,
                invoice_id=best_match.id,
                match_type=best_match_type,
                po_amount=po.amount,
                invoice_amount=best_match.amount,
                amount_diff=best_diff,
                is_exception=is_exc,
                exception_type=EXCEPTION_OVER_TOLERANCE if is_exc else None,
                status=RESULT_STATUS_PENDING,
                rule_version=rule_version,
            )
            db.session.add(mr)
            db.session.flush()
            if is_exc:
                has_exceptions = True
                ei = ExceptionItem(
                    batch_id=batch.id,
                    match_result_id=mr.id,
                    exception_type=EXCEPTION_OVER_TOLERANCE,
                    detail=f"采购单{po.po_number}与发票{best_match.invoice_number}金额差异{best_diff:.2f}，"
                           f"采购金额{po.amount}，发票金额{best_match.amount}",
                    status=EXCEPTION_STATUS_PENDING,
                )
                db.session.add(ei)
        else:
            remaining_invs = [inv for inv in vendor_invs if inv.id not in matched_inv_ids]
            fallback_inv = None
            fallback_diff = float("inf")
            for inv in remaining_invs:
                diff = abs(po.amount - inv.amount)
                if diff < fallback_diff:
                    fallback_inv = inv
                    fallback_diff = diff
            if fallback_inv is not None:
                matched_inv_ids.add(fallback_inv.id)
                mr = MatchResult(
                    batch_id=batch.id,
                    po_id=po.id,
                    invoice_id=fallback_inv.id,
                    match_type=MATCH_TYPE_OVER_TOLERANCE,
                    po_amount=po.amount,
                    invoice_amount=fallback_inv.amount,
                    amount_diff=fallback_diff,
                    is_exception=True,
                    exception_type=EXCEPTION_OVER_TOLERANCE,
                    status=RESULT_STATUS_PENDING,
                    rule_version=rule_version,
                )
                db.session.add(mr)
                db.session.flush()
                has_exceptions = True
                ei = ExceptionItem(
                    batch_id=batch.id,
                    match_result_id=mr.id,
                    exception_type=EXCEPTION_OVER_TOLERANCE,
                    detail=f"采购单{po.po_number}与发票{fallback_inv.invoice_number}金额差异{fallback_diff:.2f}，"
                           f"超出容差（采购金额{po.amount}，发票金额{fallback_inv.amount}）",
                    status=EXCEPTION_STATUS_PENDING,
                )
                db.session.add(ei)
            else:
                mr = MatchResult(
                    batch_id=batch.id,
                    po_id=po.id,
                    invoice_id=None,
                    match_type=MATCH_TYPE_UNMATCHED_PO,
                    po_amount=po.amount,
                    invoice_amount=None,
                    amount_diff=None,
                    is_exception=True,
                    exception_type=EXCEPTION_OVER_TOLERANCE,
                    status=RESULT_STATUS_PENDING,
                    rule_version=rule_version,
                )
                db.session.add(mr)
                db.session.flush()
                has_exceptions = True
                ei = ExceptionItem(
                    batch_id=batch.id,
                    match_result_id=mr.id,
                    exception_type=EXCEPTION_OVER_TOLERANCE,
                    detail=f"采购单{po.po_number}(供应商{po.vendor_code})无匹配发票",
                    status=EXCEPTION_STATUS_PENDING,
                )
                db.session.add(ei)

    for inv in invoices:
        if inv.id not in matched_inv_ids:
            mr = MatchResult(
                batch_id=batch.id,
                po_id=None,
                invoice_id=inv.id,
                match_type=MATCH_TYPE_UNMATCHED_INVOICE,
                po_amount=None,
                invoice_amount=inv.amount,
                amount_diff=None,
                is_exception=True,
                exception_type=EXCEPTION_OVER_TOLERANCE,
                status=RESULT_STATUS_PENDING,
                rule_version=rule_version,
            )
            db.session.add(mr)
            db.session.flush()
            has_exceptions = True
            ei = ExceptionItem(
                batch_id=batch.id,
                match_result_id=mr.id,
                exception_type=EXCEPTION_OVER_TOLERANCE,
                detail=f"发票{inv.invoice_number}(供应商{inv.vendor_code})无匹配采购单",
                status=EXCEPTION_STATUS_PENDING,
            )
            db.session.add(ei)

    return has_exceptions


def process_batch(batch_id):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    if not batch.can_transition(BATCH_STATUS_VALIDATING):
        raise ValidationError([f"批次状态'{batch.status}'不允许执行匹配操作"])

    batch.status = BATCH_STATUS_VALIDATING
    db.session.flush()

    all_errors = []
    try:
        if not batch.purchase_orders:
            all_errors.append("未上传采购单文件")
        if not batch.invoices:
            all_errors.append("未上传发票文件")

        if all_errors:
            raise ValidationError(all_errors)

        has_exceptions = perform_matching(batch)

        log = AuditLog(
            batch_id=batch.id,
            action="MATCH",
            detail=f"匹配完成，规则版本{batch.rule_version}",
        )
        db.session.add(log)

        if has_exceptions:
            batch.status = BATCH_STATUS_EXCEPTION
        else:
            batch.status = BATCH_STATUS_MATCHED

        db.session.commit()

        generate_payable_recalc_note(batch.id, change_source="MATCH", operator="system")

        return {"success": True, "has_exceptions": has_exceptions}

    except ValidationError:
        batch.status = BATCH_STATUS_FAILED
        log = AuditLog(batch_id=batch.id, action="MATCH_FAILED", detail="; ".join(all_errors))
        db.session.add(log)
        db.session.commit()
        raise
    except Exception as e:
        db.session.rollback()
        batch = Batch.query.get(batch_id)
        batch.status = BATCH_STATUS_FAILED
        log = AuditLog(batch_id=batch.id, action="MATCH_FAILED", detail=str(e))
        db.session.add(log)
        db.session.commit()
        raise ValidationError([str(e)])


def validate_and_import_po(batch_or_id, file_storage_or_content, filename=None):
    batch = batch_or_id if isinstance(batch_or_id, Batch) else Batch.query.get(batch_or_id)
    if isinstance(file_storage_or_content, str):
        file_storage = (file_storage_or_content, filename or batch.po_filename or "po.csv")
    else:
        file_storage = file_storage_or_content
    _, fname, _ = _extract_storage(file_storage, fallback_filename=filename)
    headers, rows = parse_file(file_storage, filename=filename)
    missing = validate_columns(headers, PO_REQUIRED_COLUMNS)
    if missing:
        raise ValidationError([f"采购单文件缺少列: {', '.join(missing)}"])

    row_errors = []
    for i, row in enumerate(rows, 2):
        row_errors.extend(validate_po_row(row, i))

    if row_errors:
        raise ValidationError(row_errors)

    PurchaseOrder.query.filter_by(batch_id=batch.id).delete()
    import_purchase_orders(batch, rows)
    batch.po_filename = fname
    log = AuditLog(batch_id=batch.id, action="UPLOAD_PO", detail=f"上传采购单 {fname}，共{len(rows)}条")
    db.session.add(log)
    db.session.commit()
    return len(rows)


def validate_and_import_invoice(batch_or_id, file_storage_or_content, filename=None):
    batch = batch_or_id if isinstance(batch_or_id, Batch) else Batch.query.get(batch_or_id)
    if isinstance(file_storage_or_content, str):
        file_storage = (file_storage_or_content, filename or batch.invoice_filename or "inv.csv")
    else:
        file_storage = file_storage_or_content
    _, fname, _ = _extract_storage(file_storage, fallback_filename=filename)
    headers, rows = parse_file(file_storage, filename=filename)
    missing = validate_columns(headers, INVOICE_REQUIRED_COLUMNS)
    if missing:
        raise ValidationError([f"发票文件缺少列: {', '.join(missing)}"])

    row_errors = []
    for i, row in enumerate(rows, 2):
        row_errors.extend(validate_invoice_row(row, i))

    dup_errors = check_duplicate_invoices(rows)
    row_errors.extend(dup_errors)

    if row_errors:
        raise ValidationError(row_errors)

    Invoice.query.filter_by(batch_id=batch.id).delete()
    import_invoices(batch, rows)
    batch.invoice_filename = fname
    log = AuditLog(batch_id=batch.id, action="UPLOAD_INVOICE", detail=f"上传发票 {fname}，共{len(rows)}条")
    db.session.add(log)
    db.session.commit()
    return len(rows)


def _compute_payable_total(batch):
    matched = [r for r in batch.match_results if r.match_type in (
        MATCH_TYPE_EXACT, MATCH_TYPE_TOLERANCE, MATCH_TYPE_OVER_TOLERANCE)]
    return round(sum(r.invoice_amount or 0 for r in matched if r.status != RESULT_STATUS_REJECTED), 2)


def _build_result_snapshot(batch):
    """构建匹配结果快照，用于版本间差异对比"""
    mr_snap = {}
    for mr in batch.match_results:
        mr_snap[str(mr.id)] = {
            "po_id": mr.po_id,
            "invoice_id": mr.invoice_id,
            "po_number": mr.po.po_number if mr.po else None,
            "invoice_number": mr.invoice.invoice_number if mr.invoice else None,
            "match_type": mr.match_type,
            "status": mr.status,
            "exception_type": mr.exception_type,
            "remarks": mr.remarks or "",
            "rule_version": mr.rule_version,
            "is_exception": mr.is_exception,
        }
    exc_snap = {}
    for exc in batch.exception_items:
        exc_snap[str(exc.id)] = {
            "match_result_id": exc.match_result_id,
            "exception_type": exc.exception_type,
            "status": exc.status,
            "remarks": exc.remarks or "",
        }
    return {
        "rule_version": batch.rule_version,
        "batch_status": batch.status,
        "match_results": mr_snap,
        "exceptions": exc_snap,
    }


def _collect_affected_documents(batch, prev_note):
    """收集当前版本相对于上一版涉及变化的采购单和发票号。
    - 首次生成（prev_note is None）：返回整批所有单据
    - 后续版本：只返回相对上一版有变化的单据
    """
    po_set = set()
    inv_set = set()

    if prev_note is None or not prev_note.result_snapshot:
        for mr in batch.match_results:
            if mr.po:
                po_set.add(mr.po.po_number)
            if mr.invoice:
                inv_set.add(mr.invoice.invoice_number)
        return sorted(po_set), sorted(inv_set)

    try:
        prev_snap = json.loads(prev_note.result_snapshot)
    except (json.JSONDecodeError, TypeError):
        for mr in batch.match_results:
            if mr.po:
                po_set.add(mr.po.po_number)
            if mr.invoice:
                inv_set.add(mr.invoice.invoice_number)
        return sorted(po_set), sorted(inv_set)

    curr_snap = _build_result_snapshot(batch)
    prev_mr = prev_snap.get("match_results", {})
    curr_mr = curr_snap.get("match_results", {})

    all_mr_ids = set(prev_mr.keys()) | set(curr_mr.keys())
    for mr_id in all_mr_ids:
        prev_item = prev_mr.get(mr_id)
        curr_item = curr_mr.get(mr_id)
        if prev_item != curr_item:
            item = curr_item or prev_item
            if item.get("po_number"):
                po_set.add(item["po_number"])
            if item.get("invoice_number"):
                inv_set.add(item["invoice_number"])

    prev_exc = prev_snap.get("exceptions", {})
    curr_exc = curr_snap.get("exceptions", {})
    all_exc_ids = set(prev_exc.keys()) | set(curr_exc.keys())
    for exc_id in all_exc_ids:
        prev_item = prev_exc.get(exc_id)
        curr_item = curr_exc.get(exc_id)
        if prev_item != curr_item:
            item = curr_item or prev_item
            mr_id = str(item.get("match_result_id")) if item.get("match_result_id") else None
            if mr_id and mr_id in curr_mr:
                if curr_mr[mr_id].get("po_number"):
                    po_set.add(curr_mr[mr_id]["po_number"])
                if curr_mr[mr_id].get("invoice_number"):
                    inv_set.add(curr_mr[mr_id]["invoice_number"])
            elif mr_id and mr_id in prev_mr:
                if prev_mr[mr_id].get("po_number"):
                    po_set.add(prev_mr[mr_id]["po_number"])
                if prev_mr[mr_id].get("invoice_number"):
                    inv_set.add(prev_mr[mr_id]["invoice_number"])

    return sorted(po_set), sorted(inv_set)


def _build_change_summary(batch, prev_note, current_total):
    """生成变化摘要文本"""
    parts = []
    if prev_note is None:
        parts.append(f"首次生成应付说明，应付合计 {current_total:.2f}")
    else:
        diff = round(current_total - prev_note.current_total, 2)
        if diff != 0:
            direction = "增加" if diff > 0 else "减少"
            parts.append(f"应付合计{direction} {abs(diff):.2f}（{prev_note.current_total:.2f} → {current_total:.2f}）")
        else:
            parts.append(f"应付合计不变（{current_total:.2f}）")
        if prev_note.rule_version != batch.rule_version:
            parts.append(f"规则版本变更: {prev_note.rule_version[:8]} → {batch.rule_version[:8]}")
    if not parts:
        parts.append("匹配结果状态或异常处理意见变更")
    return "; ".join(parts)


def generate_payable_recalc_note(batch_id, change_source=None, operator="system"):
    """
    生成应付重算说明。
    - 如果内容哈希与最新版本一致，不生成新记录（去重）
    - 否则生成新版本，并将变化摘要写入操作日志
    - 返回 (note_obj, is_new)
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        return None, False

    content_hash = compute_note_content_hash(batch)

    latest_note = (
        PayableRecalcNote.query.filter_by(batch_id=batch_id)
        .order_by(PayableRecalcNote.version.desc())
        .first()
    )

    if latest_note and latest_note.content_hash == content_hash:
        return latest_note, False

    current_total = _compute_payable_total(batch)
    previous_total = latest_note.current_total if latest_note else None
    amount_diff = round(current_total - previous_total, 2) if previous_total is not None else None

    po_numbers, invoice_numbers = _collect_affected_documents(batch, latest_note)
    change_summary = _build_change_summary(batch, latest_note, current_total)

    new_version = (latest_note.version + 1) if latest_note else 1
    snapshot = _build_result_snapshot(batch)

    note = PayableRecalcNote(
        batch_id=batch_id,
        version=new_version,
        current_total=current_total,
        previous_total=previous_total,
        amount_diff=amount_diff,
        change_source=change_source,
        change_summary=change_summary,
        po_numbers=json.dumps(po_numbers, ensure_ascii=False),
        invoice_numbers=json.dumps(invoice_numbers, ensure_ascii=False),
        rule_version=batch.rule_version,
        content_hash=content_hash,
        result_snapshot=json.dumps(snapshot, ensure_ascii=False),
    )
    db.session.add(note)

    action = "RECALC_NOTE_V1" if new_version == 1 else f"RECALC_NOTE_V{new_version}"
    log = AuditLog(
        batch_id=batch_id,
        action=action,
        detail=f"应付重算说明 v{new_version}: {change_summary}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    return note, True


def get_latest_recalc_note(batch_id):
    """获取指定批次最新的应付重算说明"""
    return (
        PayableRecalcNote.query.filter_by(batch_id=batch_id)
        .order_by(PayableRecalcNote.version.desc())
        .first()
    )


def list_recalc_notes(batch_id):
    """列出指定批次所有应付重算说明（按版本升序）"""
    return (
        PayableRecalcNote.query.filter_by(batch_id=batch_id)
        .order_by(PayableRecalcNote.version.asc())
        .all()
    )


def _get_doc_sets_from_snapshot(snapshot):
    """从快照中提取采购单和发票的集合（用于对比增删）"""
    po_set = set()
    inv_set = set()
    mr_map = {}
    exc_map = {}
    mr_id_to_docs = {}

    if not snapshot:
        return po_set, inv_set, mr_map, exc_map, mr_id_to_docs

    try:
        snap = json.loads(snapshot) if isinstance(snapshot, str) else snapshot
    except (json.JSONDecodeError, TypeError):
        return po_set, inv_set, mr_map, exc_map, mr_id_to_docs

    for mr_id, mr in snap.get("match_results", {}).items():
        po_num = mr.get("po_number")
        inv_num = mr.get("invoice_number")
        if po_num:
            po_set.add(po_num)
        if inv_num:
            inv_set.add(inv_num)
        key = f"{po_num}:{inv_num}"
        mr_map[key] = mr
        mr_id_to_docs[str(mr.get("po_id") or "")] = (po_num, inv_num)
        mr_id_to_docs[str(mr.get("invoice_id") or "")] = (po_num, inv_num)
        mr_id_to_docs[str(mr_id)] = (po_num, inv_num)

    for exc_id, exc in snap.get("exceptions", {}).items():
        exc_map[exc_id] = exc

    return po_set, inv_set, mr_map, exc_map, mr_id_to_docs


def _find_changed_docs(a_map, b_map, a_exc_map, b_exc_map, a_mr_id_to_docs, b_mr_id_to_docs, doc_type="po"):
    """找出在两个版本中都存在但状态/备注/类型等发生变化的单据"""
    changed = set()
    num_key = "po_number" if doc_type == "po" else "invoice_number"

    for key in a_map:
        if key in b_map:
            a_item = a_map[key]
            b_item = b_map[key]
            if a_item != b_item:
                num = a_item.get(num_key) or b_item.get(num_key)
                if num:
                    changed.add(num)

    all_exc_ids = set(a_exc_map.keys()) | set(b_exc_map.keys())
    for exc_id in all_exc_ids:
        a_exc = a_exc_map.get(exc_id)
        b_exc = b_exc_map.get(exc_id)
        if a_exc != b_exc:
            exc = b_exc or a_exc
            mr_id = str(exc.get("match_result_id") or "")
            docs = a_mr_id_to_docs.get(mr_id) or b_mr_id_to_docs.get(mr_id)
            if docs:
                po_num, inv_num = docs
                if doc_type == "po" and po_num:
                    changed.add(po_num)
                elif doc_type == "invoice" and inv_num:
                    changed.add(inv_num)

    return sorted(changed)


def compare_notes(batch_id, note_a_id, note_b_id, operator="system"):
    """
    比较同一批次内两个应付重算说明版本的差异。
    
    返回 (comparison_obj, error_message)。
    如果有错误，error_message 不为 None，comparison_obj 为 None。
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        return None, "批次不存在"

    note_a = PayableRecalcNote.query.get(note_a_id)
    if not note_a:
        return None, f"版本 {note_a_id} 不存在"

    note_b = PayableRecalcNote.query.get(note_b_id)
    if not note_b:
        return None, f"版本 {note_b_id} 不存在"

    if note_a.batch_id != batch_id:
        return None, f"版本 {note_a_id} 不属于批次 {batch_id}"

    if note_b.batch_id != batch_id:
        return None, f"版本 {note_b_id} 不属于批次 {batch_id}"

    if note_a_id == note_b_id:
        return None, "不能对比同一版本"

    a_po, a_inv, a_map, a_exc_map, a_mr_id_to_docs = _get_doc_sets_from_snapshot(note_a.result_snapshot)
    b_po, b_inv, b_map, b_exc_map, b_mr_id_to_docs = _get_doc_sets_from_snapshot(note_b.result_snapshot)

    po_added = sorted(b_po - a_po)
    po_removed = sorted(a_po - b_po)
    po_changed = _find_changed_docs(a_map, b_map, a_exc_map, b_exc_map, a_mr_id_to_docs, b_mr_id_to_docs, doc_type="po")

    invoice_added = sorted(b_inv - a_inv)
    invoice_removed = sorted(a_inv - b_inv)
    invoice_changed = _find_changed_docs(a_map, b_map, a_exc_map, b_exc_map, a_mr_id_to_docs, b_mr_id_to_docs, doc_type="invoice")

    amount_diff = round(note_b.current_total - note_a.current_total, 2)

    change_sources = []
    if po_added:
        change_sources.append(f"采购单新增{len(po_added)}条")
    if po_removed:
        change_sources.append(f"采购单移除{len(po_removed)}条")
    if po_changed:
        change_sources.append(f"采购单变更{len(po_changed)}条")
    if invoice_added:
        change_sources.append(f"发票新增{len(invoice_added)}条")
    if invoice_removed:
        change_sources.append(f"发票移除{len(invoice_removed)}条")
    if invoice_changed:
        change_sources.append(f"发票变更{len(invoice_changed)}条")
    if note_a.rule_version != note_b.rule_version:
        change_sources.append(f"规则版本变更")
    change_source = "; ".join(change_sources) if change_sources else "无变化"

    summary_parts = []
    if amount_diff > 0:
        summary_parts.append(f"应付合计增加 {abs(amount_diff):.2f}")
    elif amount_diff < 0:
        summary_parts.append(f"应付合计减少 {abs(amount_diff):.2f}")
    else:
        summary_parts.append("应付合计不变")
    summary_parts.append(f"(v{note_a.version} → v{note_b.version})")
    if change_sources:
        summary_parts.append(f"变化来源: {change_source}")
    comparison_summary = " ".join(summary_parts)

    detail = json.dumps({
        "note_a": {
            "version": note_a.version,
            "current_total": note_a.current_total,
            "rule_version": note_a.rule_version,
            "change_source": note_a.change_source,
        },
        "note_b": {
            "version": note_b.version,
            "current_total": note_b.current_total,
            "rule_version": note_b.rule_version,
            "change_source": note_b.change_source,
        },
        "diff": {
            "po_added": po_added,
            "po_removed": po_removed,
            "po_changed": po_changed,
            "invoice_added": invoice_added,
            "invoice_removed": invoice_removed,
            "invoice_changed": invoice_changed,
        },
    }, ensure_ascii=False)

    comparison = NoteComparison(
        batch_id=batch_id,
        note_a_id=note_a_id,
        note_b_id=note_b_id,
        note_a_version=note_a.version,
        note_b_version=note_b.version,
        amount_diff=amount_diff,
        change_source=change_source,
        po_added=json.dumps(po_added, ensure_ascii=False),
        po_removed=json.dumps(po_removed, ensure_ascii=False),
        po_changed=json.dumps(po_changed, ensure_ascii=False),
        invoice_added=json.dumps(invoice_added, ensure_ascii=False),
        invoice_removed=json.dumps(invoice_removed, ensure_ascii=False),
        invoice_changed=json.dumps(invoice_changed, ensure_ascii=False),
        rule_version_a=note_a.rule_version,
        rule_version_b=note_b.rule_version,
        operator=operator,
        comparison_summary=comparison_summary,
        detail=detail,
    )
    db.session.add(comparison)

    log = AuditLog(
        batch_id=batch_id,
        action="COMPARE_NOTES",
        detail=f"对比应付说明 v{note_a.version} vs v{note_b.version}: {comparison_summary}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return comparison, None


def get_latest_comparison(batch_id):
    """获取指定批次最新的版本对比结果"""
    return (
        NoteComparison.query.filter_by(batch_id=batch_id)
        .order_by(NoteComparison.created_at.desc())
        .first()
    )


def list_comparisons(batch_id):
    """列出指定批次所有版本对比结果（按时间降序）"""
    return (
        NoteComparison.query.filter_by(batch_id=batch_id)
        .order_by(NoteComparison.created_at.desc())
        .all()
    )


def get_comparison(comparison_id):
    """按ID获取版本对比结果"""
    return NoteComparison.query.get(comparison_id)


def list_comparisons_with_filter(batch_id, review_status=None):
    """列出指定批次的版本对比结果，支持按复核状态筛选"""
    query = NoteComparison.query.filter_by(batch_id=batch_id)
    if review_status:
        query = query.filter_by(review_status=review_status)
    return query.order_by(NoteComparison.created_at.desc()).all()


def get_latest_confirmed_comparison(batch_id):
    """获取指定批次最近一次已确认的版本对比"""
    return (
        NoteComparison.query.filter_by(batch_id=batch_id, review_status=REVIEW_STATUS_CONFIRMED)
        .order_by(NoteComparison.reviewed_at.desc())
        .first()
    )


def update_comparison_review(comparison_id, review_status, review_remark=None, operator="user"):
    """
    更新对比记录的复核状态和备注。

    返回 (comparison_obj, error_message)。
    冲突场景：
    - 对比记录不存在 → 400
    - 已确认后再确认 → 400（重复确认）
    - 已忽略后再确认 → 400（需先恢复待复核？这里按需求直接返回冲突）
    """
    comparison = NoteComparison.query.get(comparison_id)
    if not comparison:
        return None, "对比记录不存在"

    if review_status == REVIEW_STATUS_CONFIRMED:
        if comparison.review_status == REVIEW_STATUS_CONFIRMED:
            return None, "该对比记录已确认，不允许重复确认"
        if comparison.review_status == REVIEW_STATUS_IGNORED:
            return None, "该对比记录已忽略，不允许直接确认"
    if review_status == REVIEW_STATUS_IGNORED:
        if comparison.review_status == REVIEW_STATUS_CONFIRMED:
            return None, "该对比记录已确认，不允许直接忽略"
        if comparison.review_status == REVIEW_STATUS_IGNORED:
            return None, "该对比记录已忽略，不允许重复忽略"

    comparison.review_status = review_status
    comparison.review_remark = review_remark
    comparison.reviewed_by = operator
    comparison.reviewed_at = datetime.now(timezone.utc)

    status_label = {
        REVIEW_STATUS_PENDING: "待复核",
        REVIEW_STATUS_CONFIRMED: "已确认",
        REVIEW_STATUS_IGNORED: "已忽略",
    }.get(review_status, review_status)

    log = AuditLog(
        batch_id=comparison.batch_id,
        action=f"REVIEW_COMPARISON_{review_status}",
        detail=f"对比记录 #{comparison_id} 复核状态更新为 {status_label}"
               + (f"，备注: {review_remark}" if review_remark else ""),
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return comparison, None


def batch_update_comparison_review(batch_id, comparison_ids, review_status, review_remark=None, operator="user"):
    """
    批量更新对比记录的复核状态和备注。

    对每条记录做完整的冲突校验，成功的提交、冲突的逐条返回原因，
    绝不静默吞掉半成功的结果。

    返回 dict:
    {
        "success_count": int,
        "success_ids": [id, ...],
        "conflict_count": int,
        "conflicts": [{"id": int, "reason": str}, ...]
    }
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        return {
            "success_count": 0,
            "success_ids": [],
            "conflict_count": len(comparison_ids),
            "conflicts": [{"id": cid, "reason": "批次不存在"} for cid in comparison_ids],
        }

    valid_statuses = {REVIEW_STATUS_PENDING, REVIEW_STATUS_CONFIRMED, REVIEW_STATUS_IGNORED}
    if review_status not in valid_statuses:
        return {
            "success_count": 0,
            "success_ids": [],
            "conflict_count": len(comparison_ids),
            "conflicts": [{"id": cid, "reason": f"无效的复核状态: {review_status}"} for cid in comparison_ids],
        }

    status_label = {
        REVIEW_STATUS_PENDING: "待复核",
        REVIEW_STATUS_CONFIRMED: "已确认",
        REVIEW_STATUS_IGNORED: "已忽略",
    }.get(review_status, review_status)

    success_ids = []
    conflicts = []

    for cid in comparison_ids:
        comparison = NoteComparison.query.get(cid)
        if not comparison:
            conflicts.append({"id": cid, "reason": "对比记录不存在"})
            continue
        if comparison.batch_id != batch_id:
            conflicts.append({"id": cid, "reason": "对比记录不属于该批次"})
            continue
        if review_status == REVIEW_STATUS_CONFIRMED:
            if comparison.review_status == REVIEW_STATUS_CONFIRMED:
                conflicts.append({"id": cid, "reason": "该对比记录已确认，不允许重复确认"})
                continue
            if comparison.review_status == REVIEW_STATUS_IGNORED:
                conflicts.append({"id": cid, "reason": "该对比记录已忽略，不允许直接确认"})
                continue
        if review_status == REVIEW_STATUS_IGNORED:
            if comparison.review_status == REVIEW_STATUS_CONFIRMED:
                conflicts.append({"id": cid, "reason": "该对比记录已确认，不允许直接忽略"})
                continue
            if comparison.review_status == REVIEW_STATUS_IGNORED:
                conflicts.append({"id": cid, "reason": "该对比记录已忽略，不允许重复忽略"})
                continue

        comparison.review_status = review_status
        comparison.review_remark = review_remark
        comparison.reviewed_by = operator
        comparison.reviewed_at = datetime.now(timezone.utc)

        log = AuditLog(
            batch_id=batch_id,
            action=f"REVIEW_COMPARISON_{review_status}",
            detail=f"[批量] 对比记录 #{cid} 复核状态更新为 {status_label}"
                   + (f"，备注: {review_remark}" if review_remark else ""),
            operator=operator,
        )
        db.session.add(log)
        success_ids.append(cid)

    db.session.commit()

    return {
        "success_count": len(success_ids),
        "success_ids": success_ids,
        "conflict_count": len(conflicts),
        "conflicts": conflicts,
    }


def _compute_file_hash(content):
    if isinstance(content, str):
        content = content.encode("utf-8")
    return hashlib.sha256(content).hexdigest()


def check_vendor_consistency(rows, file_type):
    """检查供应商是否一致"""
    vendor_codes = set()
    vendor_names = set()
    warnings = []

    for row in rows:
        code = row.get("vendor_code", "").strip()
        name = row.get("vendor_name", "").strip()
        if code:
            vendor_codes.add(code)
        if name:
            vendor_names.add(name)

    if len(vendor_codes) > 1:
        warnings.append(
            f"文件包含 {len(vendor_codes)} 个不同的供应商编码: {', '.join(sorted(vendor_codes))}"
        )
    if len(vendor_names) > 1:
        warnings.append(
            f"文件包含 {len(vendor_names)} 个不同的供应商名称: {', '.join(sorted(vendor_names))}"
        )

    return warnings


def check_duplicate_po_numbers(rows):
    """检查采购单号重复"""
    seen = {}
    duplicates = []
    for i, row in enumerate(rows, 2):
        po_num = row.get("po_number", "").strip()
        if not po_num:
            continue
        if po_num in seen:
            duplicates.append(
                f"采购单号重复: '{po_num}' 出现在第{seen[po_num]}行和第{i}行"
            )
        else:
            seen[po_num] = i
    return duplicates


def precheck_file(file_content, filename, file_type, batch=None):
    """
    预检文件，返回预检报告。
    不写入数据库，只做解析和校验。
    """
    required_columns = PO_REQUIRED_COLUMNS if file_type == DRAFT_FILE_TYPE_PO else INVOICE_REQUIRED_COLUMNS
    validate_row_fn = validate_po_row if file_type == DRAFT_FILE_TYPE_PO else validate_invoice_row

    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig")

    headers, rows = parse_file((file_content, filename), filename=filename)

    issues = []
    report = {
        "filename": filename,
        "file_type": file_type,
        "row_count": len(rows),
        "headers": headers,
        "summary": {
            "error_count": 0,
            "warning_count": 0,
            "info_count": 0,
            "valid_rows": 0,
            "invalid_rows": 0,
        },
        "missing_columns": [],
        "issues": [],
    }

    missing = validate_columns(headers, required_columns)
    if missing:
        report["missing_columns"] = missing
        for col in missing:
            issues.append({
                "type": PRECHECK_ERROR,
                "code": "MISSING_COLUMN",
                "row_number": None,
                "column_name": col,
                "message": f"缺少必填列: {col}",
                "detail": None,
            })

    has_row_errors = False
    for i, row in enumerate(rows, 2):
        row_errors = validate_row_fn(row, i)
        for err in row_errors:
            issues.append({
                "type": PRECHECK_ERROR,
                "code": "INVALID_ROW",
                "row_number": i,
                "column_name": None,
                "message": err,
                "detail": {"row": row},
            })
            has_row_errors = True

        if row.get("amount"):
            try:
                amount = float(row["amount"])
                if amount < 0:
                    issues.append({
                        "type": PRECHECK_WARNING,
                        "code": "NEGATIVE_AMOUNT",
                        "row_number": i,
                        "column_name": "amount",
                        "message": f"第{i}行: 金额为负数 {row['amount']}",
                        "detail": {"row": row},
                    })
                if amount == 0:
                    issues.append({
                        "type": PRECHECK_WARNING,
                        "code": "ZERO_AMOUNT",
                        "row_number": i,
                        "column_name": "amount",
                        "message": f"第{i}行: 金额为0",
                        "detail": {"row": row},
                    })
            except (ValueError, TypeError):
                pass

    if file_type == DRAFT_FILE_TYPE_INVOICE:
        dup_invoices = check_duplicate_invoices(rows)
        for dup in dup_invoices:
            issues.append({
                "type": PRECHECK_ERROR,
                "code": "DUPLICATE_INVOICE",
                "row_number": None,
                "column_name": "invoice_number",
                "message": dup,
                "detail": None,
            })

    if file_type == DRAFT_FILE_TYPE_PO:
        dup_pos = check_duplicate_po_numbers(rows)
        for dup in dup_pos:
            issues.append({
                "type": PRECHECK_WARNING,
                "code": "DUPLICATE_PO",
                "row_number": None,
                "column_name": "po_number",
                "message": dup,
                "detail": None,
            })

    vendor_warnings = check_vendor_consistency(rows, file_type)
    for warning in vendor_warnings:
        issues.append({
            "type": PRECHECK_WARNING,
            "code": "VENDOR_INCONSISTENT",
            "row_number": None,
            "column_name": "vendor_code",
            "message": warning,
            "detail": None,
        })

    if batch:
        existing_vendor = None
        if file_type == DRAFT_FILE_TYPE_PO and batch.invoices:
            existing_vendor = batch.invoices[0].vendor_code
        elif file_type == DRAFT_FILE_TYPE_INVOICE and batch.purchase_orders:
            existing_vendor = batch.purchase_orders[0].vendor_code

        if existing_vendor and rows:
            file_vendors = set(row.get("vendor_code", "").strip() for row in rows if row.get("vendor_code"))
            if existing_vendor not in file_vendors:
                issues.append({
                    "type": PRECHECK_WARNING,
                    "code": "VENDOR_MISMATCH",
                    "row_number": None,
                    "column_name": "vendor_code",
                    "message": f"当前批次已有对方单据的供应商为 {existing_vendor}，本文件不包含该供应商，可能导致匹配失败",
                    "detail": {
                        "existing_vendor": existing_vendor,
                        "file_vendors": list(file_vendors),
                    },
                })

    error_count = sum(1 for i in issues if i["type"] == PRECHECK_ERROR)
    warning_count = sum(1 for i in issues if i["type"] == PRECHECK_WARNING)
    info_count = sum(1 for i in issues if i["type"] == PRECHECK_INFO)

    report["summary"]["error_count"] = error_count
    report["summary"]["warning_count"] = warning_count
    report["summary"]["info_count"] = info_count
    report["summary"]["valid_rows"] = len(rows) if not has_row_errors else sum(
        1 for i, row in enumerate(rows, 2)
        if not validate_row_fn(row, i)
    )
    report["summary"]["invalid_rows"] = len(rows) - report["summary"]["valid_rows"]
    report["issues"] = issues

    return report


def create_import_draft(batch_id, file_content, filename, file_type, operator="system"):
    """
    创建导入草稿，执行预检并保存结果。
    处理旧草稿冲突：同批次同类型的旧草稿标记为丢弃。
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    if isinstance(file_content, bytes):
        content_str = file_content.decode("utf-8-sig")
    else:
        content_str = file_content

    file_hash = _compute_file_hash(content_str)

    existing_pending = ImportDraft.query.filter_by(
        batch_id=batch_id,
        file_type=file_type,
        status=DRAFT_STATUS_PENDING,
    ).first()

    conflict_info = None
    if existing_pending:
        if existing_pending.file_hash == file_hash:
            return existing_pending, False, None

        conflict_info = {
            "old_draft_id": existing_pending.id,
            "old_filename": existing_pending.filename,
            "old_created_at": existing_pending.created_at.isoformat(),
        }
        existing_pending.status = DRAFT_STATUS_DISCARDED
        db.session.flush()

        log = AuditLog(
            batch_id=batch_id,
            action="DRAFT_CONFLICT_DISCARDED",
            detail=f"同批次同类型旧草稿 #{existing_pending.id}({existing_pending.filename}) 因重新上传被自动丢弃",
            operator=operator,
        )
        db.session.add(log)

    report = precheck_file(content_str, filename, file_type, batch)

    rule_version = compute_rule_version(batch.tolerance_pct, batch.tolerance_abs)

    draft = ImportDraft(
        batch_id=batch_id,
        file_type=file_type,
        filename=filename,
        status=DRAFT_STATUS_PENDING,
        row_count=report["row_count"],
        valid_row_count=report["summary"]["valid_rows"],
        error_count=report["summary"]["error_count"],
        warning_count=report["summary"]["warning_count"],
        tolerance_pct=batch.tolerance_pct,
        tolerance_abs=batch.tolerance_abs,
        rule_version=rule_version,
        file_content=content_str,
        file_hash=file_hash,
        parsed_data=json.dumps(report["issues"], ensure_ascii=False),
        precheck_report=json.dumps(report, ensure_ascii=False),
        operator=operator,
    )
    db.session.add(draft)
    db.session.flush()

    for issue in report["issues"]:
        db_issue = ImportDraftIssue(
            draft_id=draft.id,
            issue_type=issue["type"],
            issue_code=issue["code"],
            row_number=issue["row_number"],
            column_name=issue["column_name"],
            message=issue["message"],
            detail=json.dumps(issue["detail"], ensure_ascii=False) if issue["detail"] else None,
        )
        db.session.add(db_issue)

    log_action = "DRAFT_CREATED_PO" if file_type == DRAFT_FILE_TYPE_PO else "DRAFT_CREATED_INVOICE"
    log_detail = (
        f"创建{'采购单' if file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿: {filename}, "
        f"共{report['row_count']}行, "
        f"错误{report['summary']['error_count']}个, "
        f"警告{report['summary']['warning_count']}个"
    )
    if conflict_info:
        log_detail += f" (冲突: 旧草稿 #{conflict_info['old_draft_id']} 被丢弃)"

    log = AuditLog(
        batch_id=batch_id,
        action=log_action,
        detail=log_detail,
        operator=operator,
    )
    db.session.add(log)

    db.session.commit()
    db.session.refresh(draft)

    return draft, True, conflict_info


def get_latest_draft(batch_id, file_type=None):
    """获取最新草稿"""
    query = ImportDraft.query.filter_by(batch_id=batch_id)
    if file_type:
        query = query.filter_by(file_type=file_type)
    return query.order_by(ImportDraft.created_at.desc()).first()


def list_drafts(batch_id, file_type=None, status=None):
    """列出草稿"""
    query = ImportDraft.query.filter_by(batch_id=batch_id)
    if file_type:
        query = query.filter_by(file_type=file_type)
    if status:
        query = query.filter_by(status=status)
    return query.order_by(ImportDraft.created_at.desc()).all()


def get_draft(draft_id):
    """获取单个草稿"""
    return ImportDraft.query.get(draft_id)


def confirm_draft(draft_id, operator="system"):
    """
    确认草稿，将数据写入正式表。
    确认后草稿状态变为 CONFIRMED。
    """
    draft = ImportDraft.query.get(draft_id)
    if not draft:
        raise ValidationError(["草稿不存在"])

    if draft.status != DRAFT_STATUS_PENDING:
        raise ValidationError([f"草稿状态为 '{draft.status}'，不允许确认"])

    batch = Batch.query.get(draft.batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    try:
        if draft.file_type == DRAFT_FILE_TYPE_PO:
            count = validate_and_import_po(batch.id, draft.file_content, draft.filename)
        else:
            count = validate_and_import_invoice(batch.id, draft.file_content, draft.filename)

        draft.status = DRAFT_STATUS_CONFIRMED
        db.session.flush()

        log_action = "DRAFT_CONFIRMED_PO" if draft.file_type == DRAFT_FILE_TYPE_PO else "DRAFT_CONFIRMED_INVOICE"
        log = AuditLog(
            batch_id=batch.id,
            action=log_action,
            detail=f"确认{'采购单' if draft.file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿 #{draft_id}，写入 {count} 行数据",
            operator=operator,
        )
        db.session.add(log)

        db.session.commit()

        return {
            "success": True,
            "imported_count": count,
            "draft_id": draft_id,
        }

    except ValidationError:
        db.session.rollback()
        raise
    except Exception as e:
        db.session.rollback()
        raise ValidationError([str(e)])


def discard_draft(draft_id, operator="system"):
    """
    丢弃草稿，不写入数据。
    """
    draft = ImportDraft.query.get(draft_id)
    if not draft:
        raise ValidationError(["草稿不存在"])

    if draft.status != DRAFT_STATUS_PENDING:
        raise ValidationError([f"草稿状态为 '{draft.status}'，不允许丢弃"])

    draft.status = DRAFT_STATUS_DISCARDED
    db.session.flush()

    log_action = "DRAFT_DISCARDED_PO" if draft.file_type == DRAFT_FILE_TYPE_PO else "DRAFT_DISCARDED_INVOICE"
    log = AuditLog(
        batch_id=draft.batch_id,
        action=log_action,
        detail=f"丢弃{'采购单' if draft.file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿 #{draft_id} ({draft.filename})",
        operator=operator,
    )
    db.session.add(log)

    db.session.commit()

    return {
        "success": True,
        "draft_id": draft_id,
    }


def cancel_draft(draft_id, operator="system"):
    """
    取消草稿（与丢弃类似，但语义不同：用户主动取消 vs 被新草稿替换）。
    取消后原正式数据保持不变。
    """
    draft = ImportDraft.query.get(draft_id)
    if not draft:
        raise ValidationError(["草稿不存在"])

    if draft.status != DRAFT_STATUS_PENDING:
        raise ValidationError([f"草稿状态为 '{draft.status}'，不允许取消"])

    draft.status = DRAFT_STATUS_CANCELLED
    db.session.flush()

    log_action = "DRAFT_CANCELLED_PO" if draft.file_type == DRAFT_FILE_TYPE_PO else "DRAFT_CANCELLED_INVOICE"
    log = AuditLog(
        batch_id=draft.batch_id,
        action=log_action,
        detail=f"取消{'采购单' if draft.file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿 #{draft_id} ({draft.filename})，原正式数据保持不变",
        operator=operator,
    )
    db.session.add(log)

    db.session.commit()

    return {
        "success": True,
        "draft_id": draft_id,
    }


def _row_signature(row, file_type):
    """生成行的业务主键签名（用于对比覆盖/新增/跳过）。"""
    if file_type == DRAFT_FILE_TYPE_PO:
        return (
            row.get("po_number", "").strip(),
            row.get("vendor_code", "").strip(),
        )
    else:
        return (
            row.get("invoice_number", "").strip(),
            row.get("vendor_code", "").strip(),
        )


def _row_value_signature(row, file_type):
    """生成行内容签名（用于判断内容是否变化）。"""
    if file_type == DRAFT_FILE_TYPE_PO:
        return (
            round(float(row.get("amount", 0)), 2),
            (row.get("vendor_name", "") or "").strip(),
            (row.get("po_date", "") or "").strip(),
            (row.get("currency", "CNY") or "CNY").strip(),
        )
    else:
        return (
            round(float(row.get("amount", 0)), 2),
            (row.get("vendor_name", "") or "").strip(),
            (row.get("invoice_date", "") or "").strip(),
            (row.get("currency", "CNY") or "CNY").strip(),
        )


def analyze_diff(batch_id, file_type, parsed_rows, existing_draft_id=None):
    """
    分析草稿与正式数据、以及与同批次上一版草稿的差异。

    返回 dict:
    {
        "vs_official": {
            "add_count": int,
            "overwrite_count": int,
            "skip_count": int,
            "conflict_count": int,
            "add_rows": [ {row_index, key, amount, vendor} ... ],
            "overwrite_rows": [ ... ],
            "skip_rows": [ ... ],
            "conflict_rows": [ ... ],
        },
        "vs_previous_draft": {
            "prev_draft_id": int or None,
            "prev_filename": str or None,
            "same_file": bool,
            "changed_count": int,
            "added_vs_prev": [ ... ],
            "removed_vs_prev": [ ... ],
            "modified_vs_prev": [ ... ],
        },
        "cross_batch_conflicts": {
            "invoice_duplicates": [ {invoice_number, existing_batch_id, existing_batch_name, amount, vendor} ],
        },
        "summary_text": str,
    }
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    required_columns = PO_REQUIRED_COLUMNS if file_type == DRAFT_FILE_TYPE_PO else INVOICE_REQUIRED_COLUMNS
    validate_row_fn = validate_po_row if file_type == DRAFT_FILE_TYPE_PO else validate_invoice_row

    valid_rows = []
    for i, row in enumerate(parsed_rows):
        row_num = i + 2
        errs = validate_row_fn(row, row_num)
        if not errs:
            valid_rows.append((i, row))

    vs_official = {
        "add_count": 0,
        "overwrite_count": 0,
        "skip_count": 0,
        "conflict_count": 0,
        "add_rows": [],
        "overwrite_rows": [],
        "skip_rows": [],
        "conflict_rows": [],
    }

    if file_type == DRAFT_FILE_TYPE_PO:
        existing_records = PurchaseOrder.query.filter_by(batch_id=batch_id).all()
    else:
        existing_records = Invoice.query.filter_by(batch_id=batch_id).all()

    existing_map = {}
    for rec in existing_records:
        if file_type == DRAFT_FILE_TYPE_PO:
            key = (rec.po_number.strip(), rec.vendor_code.strip())
            val_sig = (round(rec.amount, 2), (rec.vendor_name or "").strip(), (rec.po_date or "").strip(), (rec.currency or "CNY").strip())
        else:
            key = (rec.invoice_number.strip(), rec.vendor_code.strip())
            val_sig = (round(rec.amount, 2), (rec.vendor_name or "").strip(), (rec.invoice_date or "").strip(), (rec.currency or "CNY").strip())
        existing_map[key] = {
            "id": rec.id,
            "value_sig": val_sig,
            "amount": rec.amount,
            "vendor": rec.vendor_name or rec.vendor_code,
        }

    seen_keys_in_draft = set()
    for i, row in valid_rows:
        key = _row_signature(row, file_type)
        if not key[0]:
            continue
        try:
            val_sig = _row_value_signature(row, file_type)
        except (ValueError, TypeError):
            continue

        row_num = i + 2
        amount = float(row.get("amount", 0) or 0)
        vendor = row.get("vendor_name") or row.get("vendor_code") or ""
        doc_number = key[0]

        if key in seen_keys_in_draft:
            vs_official["conflict_count"] += 1
            vs_official["conflict_rows"].append({
                "row_index": row_num,
                "key": doc_number,
                "amount": amount,
                "vendor": vendor,
                "reason": "草稿内重复",
            })
            continue
        seen_keys_in_draft.add(key)

        if key in existing_map:
            existing = existing_map[key]
            if val_sig == existing["value_sig"]:
                vs_official["skip_count"] += 1
                vs_official["skip_rows"].append({
                    "row_index": row_num,
                    "key": doc_number,
                    "amount": amount,
                    "vendor": vendor,
                    "existing_id": existing["id"],
                    "existing_amount": existing["amount"],
                })
            else:
                vs_official["overwrite_count"] += 1
                vs_official["overwrite_rows"].append({
                    "row_index": row_num,
                    "key": doc_number,
                    "amount": amount,
                    "vendor": vendor,
                    "existing_id": existing["id"],
                    "existing_amount": existing["amount"],
                    "diff_amount": round(amount - existing["amount"], 2),
                })
        else:
            vs_official["add_count"] += 1
            vs_official["add_rows"].append({
                "row_index": row_num,
                "key": doc_number,
                "amount": amount,
                "vendor": vendor,
            })

    vs_previous = {
        "prev_draft_id": None,
        "prev_filename": None,
        "same_file": False,
        "changed_count": 0,
        "added_vs_prev": [],
        "removed_vs_prev": [],
        "modified_vs_prev": [],
    }

    query = ImportDraft.query.filter_by(
        batch_id=batch_id,
        file_type=file_type,
    ).filter(ImportDraft.id != existing_draft_id if existing_draft_id else True)
    prev_draft = query.order_by(ImportDraft.created_at.desc()).first()
    if prev_draft and prev_draft.precheck_report:
        try:
            prev_report = json.loads(prev_draft.precheck_report)
            prev_headers = prev_report.get("headers", [])
            prev_issue_codes = set()
            for iss in prev_report.get("issues", []):
                if iss.get("type") == PRECHECK_ERROR:
                    prev_issue_codes.add(iss.get("code"))

            if "MISSING_COLUMN" not in prev_issue_codes:
                prev_content = prev_draft.file_content
                _, prev_rows = parse_file((prev_content, prev_draft.filename), filename=prev_draft.filename)

                prev_valid_keys = {}
                for i, row in enumerate(prev_rows):
                    errs = validate_row_fn(row, i + 2)
                    if errs:
                        continue
                    key = _row_signature(row, file_type)
                    if not key[0]:
                        continue
                    try:
                        vsig = _row_value_signature(row, file_type)
                    except (ValueError, TypeError):
                        continue
                    prev_valid_keys[key] = {
                        "amount": float(row.get("amount", 0) or 0),
                        "vendor": row.get("vendor_name") or row.get("vendor_code") or "",
                        "vsig": vsig,
                    }

                curr_valid_keys = {}
                for i, row in valid_rows:
                    key = _row_signature(row, file_type)
                    if not key[0]:
                        continue
                    try:
                        vsig = _row_value_signature(row, file_type)
                    except (ValueError, TypeError):
                        continue
                    curr_valid_keys[key] = {
                        "amount": float(row.get("amount", 0) or 0),
                        "vendor": row.get("vendor_name") or row.get("vendor_code") or "",
                        "vsig": vsig,
                    }

                added_keys = set(curr_valid_keys.keys()) - set(prev_valid_keys.keys())
                removed_keys = set(prev_valid_keys.keys()) - set(curr_valid_keys.keys())
                common_keys = set(curr_valid_keys.keys()) & set(prev_valid_keys.keys())

                for k in added_keys:
                    info = curr_valid_keys[k]
                    vs_previous["added_vs_prev"].append({
                        "key": k[0],
                        "amount": info["amount"],
                        "vendor": info["vendor"],
                    })
                for k in removed_keys:
                    info = prev_valid_keys[k]
                    vs_previous["removed_vs_prev"].append({
                        "key": k[0],
                        "amount": info["amount"],
                        "vendor": info["vendor"],
                    })
                for k in common_keys:
                    a = prev_valid_keys[k]
                    b = curr_valid_keys[k]
                    if a["vsig"] != b["vsig"]:
                        vs_previous["modified_vs_prev"].append({
                            "key": k[0],
                            "old_amount": a["amount"],
                            "new_amount": b["amount"],
                            "vendor": b["vendor"],
                            "diff_amount": round(b["amount"] - a["amount"], 2),
                        })

                vs_previous["prev_draft_id"] = prev_draft.id
                vs_previous["prev_filename"] = prev_draft.filename
                vs_previous["same_file"] = (prev_draft.file_hash == _compute_file_hash(parsed_rows_to_csv(parsed_rows, required_columns)))
                vs_previous["changed_count"] = (
                    len(vs_previous["added_vs_prev"])
                    + len(vs_previous["removed_vs_prev"])
                    + len(vs_previous["modified_vs_prev"])
                )
        except Exception:
            pass

    cross_batch = {
        "invoice_duplicates": [],
    }
    if file_type == DRAFT_FILE_TYPE_INVOICE:
        seen_inv_in_draft = set()
        for i, row in valid_rows:
            inv_num = row.get("invoice_number", "").strip()
            if not inv_num or inv_num in seen_inv_in_draft:
                continue
            seen_inv_in_draft.add(inv_num)
            dup = (
                Invoice.query
                .join(Batch, Invoice.batch_id == Batch.id)
                .filter(
                    Invoice.invoice_number == inv_num,
                    Invoice.batch_id != batch_id,
                )
                .with_entities(
                    Invoice.id, Invoice.invoice_number, Invoice.amount,
                    Invoice.vendor_code, Invoice.vendor_name,
                    Batch.id.label("bid"), Batch.name.label("bname"),
                )
                .first()
            )
            if dup:
                try:
                    amount = float(row.get("amount", 0) or 0)
                except (ValueError, TypeError):
                    amount = 0
                cross_batch["invoice_duplicates"].append({
                    "invoice_number": inv_num,
                    "existing_batch_id": dup.bid,
                    "existing_batch_name": dup.bname,
                    "existing_amount": float(dup.amount or 0),
                    "draft_amount": amount,
                    "vendor": row.get("vendor_name") or row.get("vendor_code") or "",
                })

    summary_parts = []
    if vs_official["add_count"]:
        summary_parts.append(f"新增 {vs_official['add_count']} 条")
    if vs_official["overwrite_count"]:
        summary_parts.append(f"覆盖 {vs_official['overwrite_count']} 条")
    if vs_official["skip_count"]:
        summary_parts.append(f"跳过 {vs_official['skip_count']} 条")
    if vs_official["conflict_count"]:
        summary_parts.append(f"冲突 {vs_official['conflict_count']} 条")
    if cross_batch["invoice_duplicates"]:
        summary_parts.append(f"跨批次重复发票 {len(cross_batch['invoice_duplicates'])} 条")
    if not summary_parts:
        summary_parts.append("无变化")
    summary_text = "；".join(summary_parts)

    return {
        "vs_official": vs_official,
        "vs_previous_draft": vs_previous,
        "cross_batch_conflicts": cross_batch,
        "summary_text": summary_text,
    }


def parsed_rows_to_csv(rows, headers):
    """将解析后的行重新序列化为 CSV 字符串用于哈希比较。"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in headers})
    return output.getvalue()


def create_import_draft(batch_id, file_content, filename, file_type, operator="system"):
    """
    创建导入草稿，执行预检、diff 分析并保存结果。
    处理旧草稿冲突：同批次同类型的旧 PENDING 草稿被标记为丢弃并记录原因。
    """
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    if isinstance(file_content, bytes):
        content_str = file_content.decode("utf-8-sig")
    else:
        content_str = file_content

    file_hash = _compute_file_hash(content_str)

    required_columns = PO_REQUIRED_COLUMNS if file_type == DRAFT_FILE_TYPE_PO else INVOICE_REQUIRED_COLUMNS
    headers, parsed_rows = parse_file((content_str, filename), filename=filename)

    existing_pending = ImportDraft.query.filter_by(
        batch_id=batch_id,
        file_type=file_type,
        status=DRAFT_STATUS_PENDING,
    ).first()

    conflict_info = None
    conflict_reason = None
    supersedes_draft_id = None

    if existing_pending:
        if existing_pending.file_hash == file_hash:
            return existing_pending, False, None

        conflict_info = {
            "old_draft_id": existing_pending.id,
            "old_filename": existing_pending.filename,
            "old_created_at": existing_pending.created_at.isoformat(),
            "reason": "同批次同类型文件重新上传，旧草稿被自动丢弃",
        }
        existing_pending.status = DRAFT_STATUS_DISCARDED
        existing_pending.conflict_reason = "同批次同类型重新上传，被新草稿取代"
        existing_pending.superseded_by_draft_id = None
        supersedes_draft_id = existing_pending.id
        db.session.flush()

        log = AuditLog(
            batch_id=batch_id,
            action="DRAFT_CONFLICT_DISCARDED",
            detail=f"同批次同类型旧草稿 #{existing_pending.id}({existing_pending.filename}) 因重新上传被自动丢弃",
            operator=operator,
        )
        db.session.add(log)

    report = precheck_file(content_str, filename, file_type, batch)

    missing_cols = validate_columns(headers, required_columns)
    has_fatal_errors = bool(missing_cols) or report["summary"]["error_count"] > 0
    cross_batch_conflicts_empty = True

    if not has_fatal_errors:
        try:
            diff = analyze_diff(batch_id, file_type, parsed_rows)
            cross_batch_conflicts_empty = len(diff["cross_batch_conflicts"]["invoice_duplicates"]) == 0
        except Exception:
            diff = None
            cross_batch_conflicts_empty = True
    else:
        diff = None

    if not has_fatal_errors and diff and not cross_batch_conflicts_empty:
        conflict_reason = f"检测到 {len(diff['cross_batch_conflicts']['invoice_duplicates'])} 条跨批次重复发票，导入前请人工复核"

    rule_version = compute_rule_version(batch.tolerance_pct, batch.tolerance_abs)

    review_summary = None
    if diff:
        review_summary = diff["summary_text"]

    draft = ImportDraft(
        batch_id=batch_id,
        file_type=file_type,
        filename=filename,
        status=DRAFT_STATUS_PENDING if cross_batch_conflicts_empty else DRAFT_STATUS_CONFLICT,
        row_count=report["row_count"],
        valid_row_count=report["summary"]["valid_rows"],
        error_count=report["summary"]["error_count"],
        warning_count=report["summary"]["warning_count"],
        tolerance_pct=batch.tolerance_pct,
        tolerance_abs=batch.tolerance_abs,
        rule_version=rule_version,
        file_content=content_str,
        file_hash=file_hash,
        parsed_data=json.dumps(report["issues"], ensure_ascii=False),
        precheck_report=json.dumps(report, ensure_ascii=False),
        diff_analysis=json.dumps(diff, ensure_ascii=False) if diff else None,
        conflict_reason=conflict_reason,
        review_summary=review_summary,
        operator=operator,
        supersedes_draft_id=supersedes_draft_id,
    )
    db.session.add(draft)
    db.session.flush()

    if supersedes_draft_id is not None:
        prev = ImportDraft.query.get(supersedes_draft_id)
        if prev:
            prev.superseded_by_draft_id = draft.id

    for issue in report["issues"]:
        db_issue = ImportDraftIssue(
            draft_id=draft.id,
            issue_type=issue["type"],
            issue_code=issue["code"],
            row_number=issue["row_number"],
            column_name=issue["column_name"],
            message=issue["message"],
            detail=json.dumps(issue["detail"], ensure_ascii=False) if issue["detail"] else None,
        )
        db.session.add(db_issue)

    log_action = "DRAFT_CREATED_PO" if file_type == DRAFT_FILE_TYPE_PO else "DRAFT_CREATED_INVOICE"
    log_detail = (
        f"创建{'采购单' if file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿: {filename}, "
        f"共{report['row_count']}行, "
        f"错误{report['summary']['error_count']}个, "
        f"警告{report['summary']['warning_count']}个"
    )
    if review_summary:
        log_detail += f" | 复核摘要: {review_summary}"
    if conflict_info:
        log_detail += f" (冲突: 旧草稿 #{conflict_info['old_draft_id']} 被丢弃)"
    if not cross_batch_conflicts_empty:
        log_detail += f" | 状态: CONFLICT ({conflict_reason})"

    log = AuditLog(
        batch_id=batch_id,
        action=log_action,
        detail=log_detail,
        operator=operator,
    )
    db.session.add(log)

    db.session.commit()
    db.session.refresh(draft)

    return draft, True, conflict_info


def confirm_draft(draft_id, operator="system"):
    """
    确认草稿，将数据写入正式表。
    阻断逻辑：
    1. 草稿状态只能是 PENDING 或 CONFLICT（CONFLICT 需显式确认，表示人工已复核）
    2. 缺列/格式错误文件不能确认
    3. 跨批次重复发票不能确认（必须人工先处理）
    4. 草稿必须属于批次，不能跨批次误操作
    """
    draft = ImportDraft.query.get(draft_id)
    if not draft:
        raise ValidationError(["草稿不存在"])

    if draft.status not in (DRAFT_STATUS_PENDING, DRAFT_STATUS_CONFLICT):
        raise ValidationError([f"草稿状态为 '{draft.status}'，不允许确认"])

    if draft.created_at and datetime.now(timezone.utc) - draft.created_at.replace(tzinfo=timezone.utc) > timedelta(hours=DRAFT_EXPIRE_HOURS):
        raise ValidationError([f"草稿已超过 {DRAFT_EXPIRE_HOURS} 小时有效期，请重新上传文件生成新草稿"])

    batch = Batch.query.get(draft.batch_id)
    if not batch:
        raise ValidationError(["关联批次不存在"])

    if draft.diff_analysis:
        try:
            diff = json.loads(draft.diff_analysis)
            cross = diff.get("cross_batch_conflicts", {})
            dups = cross.get("invoice_duplicates", [])
            if dups:
                dup_nums = ", ".join(d["invoice_number"] for d in dups[:5])
                more = "" if len(dups) <= 5 else f" 等{len(dups)}条"
                raise ValidationError([
                    f"存在跨批次重复发票: {dup_nums}{more}，"
                    f"请先在对应批次处理后再确认。使用 /discard 或 /cancel 放弃当前草稿。"
                ])
        except ValidationError:
            raise
        except Exception:
            pass

    if draft.precheck_report:
        try:
            rpt = json.loads(draft.precheck_report)
            miss = rpt.get("missing_columns", [])
            if miss:
                raise ValidationError([f"文件缺少必要列: {', '.join(miss)}，请修正文件后重新上传"])
            err_count = rpt.get("summary", {}).get("error_count", 0)
            if err_count > 0:
                raise ValidationError([f"文件存在 {err_count} 条格式错误，请修正后重新上传；可查看草稿的 issues 详情"])
        except ValidationError:
            raise
        except Exception:
            pass

    try:
        if draft.file_type == DRAFT_FILE_TYPE_PO:
            count = validate_and_import_po(batch.id, draft.file_content, draft.filename)
        else:
            count = validate_and_import_invoice(batch.id, draft.file_content, draft.filename)

        draft.status = DRAFT_STATUS_CONFIRMED
        draft.confirmed_by = operator
        draft.confirmed_at = datetime.now(timezone.utc)
        db.session.flush()

        log_action = "DRAFT_CONFIRMED_PO" if draft.file_type == DRAFT_FILE_TYPE_PO else "DRAFT_CONFIRMED_INVOICE"
        review_extra = f"；复核摘要: {draft.review_summary}" if draft.review_summary else ""
        log = AuditLog(
            batch_id=batch.id,
            action=log_action,
            detail=f"确认{'采购单' if draft.file_type == DRAFT_FILE_TYPE_PO else '发票'}草稿 #{draft_id}，写入 {count} 行数据（确认人: {operator}）{review_extra}",
            operator=operator,
        )
        db.session.add(log)

        db.session.commit()

        return {
            "success": True,
            "imported_count": count,
            "draft_id": draft_id,
            "confirmed_by": operator,
            "review_summary": draft.review_summary,
        }

    except ValidationError:
        db.session.rollback()
        raise
    except Exception as e:
        db.session.rollback()
        raise ValidationError([str(e)])


def get_latest_confirmed_draft(batch_id, file_type=None):
    """获取指定批次最近一次已确认的草稿（用于导出 CSV 复核摘要）。"""
    query = ImportDraft.query.filter_by(
        batch_id=batch_id,
        status=DRAFT_STATUS_CONFIRMED,
    )
    if file_type:
        query = query.filter_by(file_type=file_type)
    return query.order_by(ImportDraft.confirmed_at.desc()).first()


def get_latest_review_summary(batch_id):
    """获取批次最近一次预检/复核摘要（包含采购单和发票最近一次）。"""
    po_draft = get_latest_confirmed_draft(batch_id, file_type=DRAFT_FILE_TYPE_PO)
    inv_draft = get_latest_confirmed_draft(batch_id, file_type=DRAFT_FILE_TYPE_INVOICE)

    parts = []
    if po_draft:
        summary = po_draft.review_summary or f"已确认采购单草稿 #{po_draft.id}"
        parts.append(f"[采购单] {summary} @{po_draft.confirmed_by or 'system'}")
    if inv_draft:
        summary = inv_draft.review_summary or f"已确认发票草稿 #{inv_draft.id}"
        parts.append(f"[发票] {summary} @{inv_draft.confirmed_by or 'system'}")

    return "；".join(parts) if parts else None


def create_import_plan(batch_id, po_content=None, po_filename=None,
                       invoice_content=None, invoice_filename=None, operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])
    if not po_content and not invoice_content:
        raise ValidationError(["must provide at least one file"])

    plan = ImportPlan(
        batch_id=batch_id,
        status=PLAN_STATUS_PENDING,
        operator=operator,
    )
    db.session.add(plan)
    db.session.flush()

    draft_results = {}
    if po_content:
        draft, is_new, conflict = create_import_draft(
            batch_id, po_content, po_filename or "po.csv",
            DRAFT_FILE_TYPE_PO, operator=operator,
        )
        draft.plan_id = plan.id
        db.session.flush()
        draft_results["po"] = draft.to_dict()

    if invoice_content:
        draft, is_new, conflict = create_import_draft(
            batch_id, invoice_content, invoice_filename or "invoices.csv",
            DRAFT_FILE_TYPE_INVOICE, operator=operator,
        )
        draft.plan_id = plan.id
        db.session.flush()
        draft_results["invoice"] = draft.to_dict()

    summary_parts = []
    for key, d in draft_results.items():
        label = "采购单" if key == "po" else "发票"
        da = d.get("diff_analysis") or {}
        vs = da.get("vs_official") or {}
        summary_parts.append(
            f"[{label}] 新增{vs.get('add_count',0)}条;"
            f"覆盖{vs.get('overwrite_count',0)}条;"
            f"跳过{vs.get('skip_count',0)}条;"
            f"冲突{vs.get('conflict_count',0)}条"
        )
    plan.plan_summary = json.dumps({
        "parts": summary_parts,
        "po_draft_id": draft_results.get("po", {}).get("id"),
        "invoice_draft_id": draft_results.get("invoice", {}).get("id"),
    }, ensure_ascii=False)

    db.session.add(AuditLog(
        batch_id=batch_id,
        action="PLAN_CREATED",
        detail=f"create plan #{plan.id}: {'; '.join(summary_parts)}",
        operator=operator,
    ))
    db.session.commit()
    db.session.refresh(plan)
    return plan


def get_plan(plan_id):
    return ImportPlan.query.get(plan_id)


def list_plans(batch_id, status=None):
    q = ImportPlan.query.filter_by(batch_id=batch_id)
    if status:
        q = q.filter_by(status=status)
    return q.order_by(ImportPlan.created_at.desc()).all()


def get_latest_plan(batch_id):
    return ImportPlan.query.filter_by(batch_id=batch_id).order_by(ImportPlan.created_at.desc()).first()


def confirm_plan(plan_id, operator="system"):
    plan = ImportPlan.query.get(plan_id)
    if not plan:
        raise ValidationError(["plan not found"])
    if plan.status != PLAN_STATUS_PENDING:
        raise ValidationError([f"plan status '{plan.status}' not confirmable"])

    if plan.created_at and datetime.now(timezone.utc) - plan.created_at.replace(tzinfo=timezone.utc) > timedelta(hours=DRAFT_EXPIRE_HOURS):
        raise ValidationError([f"plan expired over {DRAFT_EXPIRE_HOURS}h, please re-upload"])

    drafts = ImportDraft.query.filter_by(plan_id=plan_id).all()
    for d in drafts:
        if d.created_at and datetime.now(timezone.utc) - d.created_at.replace(tzinfo=timezone.utc) > timedelta(hours=DRAFT_EXPIRE_HOURS):
            raise ValidationError([f"draft #{d.id} expired over {DRAFT_EXPIRE_HOURS}h, please re-upload"])

    for d in drafts:
        if d.status not in (DRAFT_STATUS_PENDING, DRAFT_STATUS_CONFLICT):
            raise ValidationError([f"draft #{d.id} status '{d.status}' not confirmable"])

    for d in drafts:
        if d.diff_analysis:
            try:
                diff = json.loads(d.diff_analysis)
                cross = diff.get("cross_batch_conflicts", {})
                dups = cross.get("invoice_duplicates", [])
                if dups:
                    dup_nums = ", ".join(x["invoice_number"] for x in dups[:5])
                    raise ValidationError([f"draft #{d.id} has cross-batch duplicate invoices: {dup_nums}"])
            except ValidationError:
                raise
            except Exception:
                pass
        if d.precheck_report:
            try:
                rpt = json.loads(d.precheck_report)
                miss = rpt.get("missing_columns", [])
                if miss:
                    raise ValidationError([f"draft #{d.id} missing columns: {', '.join(miss)}"])
                if rpt.get("summary", {}).get("error_count", 0) > 0:
                    raise ValidationError([f"draft #{d.id} has {rpt['summary']['error_count']} format errors"])
            except ValidationError:
                raise
            except Exception:
                pass

    po_before = {r.id: r for r in PurchaseOrder.query.filter_by(batch_id=plan.batch_id).all()}
    inv_before = {r.id: r for r in Invoice.query.filter_by(batch_id=plan.batch_id).all()}

    for d in drafts:
        _snapshot_before_confirm(d, plan.id, po_before, inv_before)

    import_results = []
    for d in drafts:
        if d.file_type == DRAFT_FILE_TYPE_PO:
            count = validate_and_import_po(plan.batch_id, d.file_content, d.filename)
        else:
            count = validate_and_import_invoice(plan.batch_id, d.file_content, d.filename)
        d.status = DRAFT_STATUS_CONFIRMED
        d.confirmed_by = operator
        d.confirmed_at = datetime.now(timezone.utc)
        import_results.append({"file_type": d.file_type, "imported_count": count})

    plan.status = PLAN_STATUS_CONFIRMED
    plan.confirmed_by = operator
    plan.confirmed_at = datetime.now(timezone.utc)

    db.session.add(AuditLog(
        batch_id=plan.batch_id,
        action="PLAN_CONFIRMED",
        detail=f"confirm plan #{plan_id}, results: {json.dumps(import_results, ensure_ascii=False)}, by {operator}",
        operator=operator,
    ))
    db.session.commit()
    return {
        "success": True,
        "plan_id": plan_id,
        "confirmed_by": operator,
        "import_results": import_results,
    }


def _snapshot_before_confirm(draft, plan_id, po_before, inv_before):
    if draft.file_type == DRAFT_FILE_TYPE_PO:
        existing = po_before
        model_cls = PurchaseOrder
    else:
        existing = inv_before
        model_cls = Invoice

    if draft.diff_analysis:
        try:
            diff = json.loads(draft.diff_analysis)
            vs = diff.get("vs_official", {})
            for row in vs.get("overwrite_rows", []):
                sig = row.get("signature", "")
                for rid, rec in existing.items():
                    if draft.file_type == DRAFT_FILE_TYPE_PO:
                        rec_sig = f"{rec.po_number}|{rec.vendor_code}"
                    else:
                        rec_sig = f"{rec.invoice_number}|{rec.vendor_code}"
                    if rec_sig == sig:
                        db.session.add(PlanSnapshot(
                            plan_id=plan_id,
                            table_name=model_cls.__tablename__,
                            row_id=rid,
                            action=ROW_ACTION_OVERWRITE,
                            original_data=json.dumps(rec.to_dict(), ensure_ascii=False),
                        ))
                        break
        except Exception:
            pass


def cancel_plan(plan_id, operator="system"):
    plan = ImportPlan.query.get(plan_id)
    if not plan:
        raise ValidationError(["plan not found"])
    if plan.status != PLAN_STATUS_PENDING:
        raise ValidationError([f"plan status '{plan.status}' not cancellable"])

    drafts = ImportDraft.query.filter_by(plan_id=plan_id).all()
    for d in drafts:
        if d.status in (DRAFT_STATUS_PENDING, DRAFT_STATUS_CONFLICT):
            d.status = DRAFT_STATUS_CANCELLED

    plan.status = PLAN_STATUS_CANCELLED
    plan.cancelled_by = operator
    plan.cancelled_at = datetime.now(timezone.utc)

    db.session.add(AuditLog(
        batch_id=plan.batch_id,
        action="PLAN_CANCELLED",
        detail=f"cancel plan #{plan_id} by {operator}, original data unchanged",
        operator=operator,
    ))
    db.session.commit()
    return {"success": True, "plan_id": plan_id, "note": "plan cancelled, original data unchanged"}


def undo_plan(plan_id, operator="system"):
    plan = ImportPlan.query.get(plan_id)
    if not plan:
        raise ValidationError(["plan not found"])
    if plan.status != PLAN_STATUS_CONFIRMED:
        raise ValidationError([f"plan status '{plan.status}' cannot be undone"])

    latest_confirmed = ImportPlan.query.filter_by(
        batch_id=plan.batch_id, status=PLAN_STATUS_CONFIRMED,
    ).order_by(ImportPlan.confirmed_at.desc()).first()
    if latest_confirmed and latest_confirmed.id != plan.id:
        raise ValidationError(["only the most recent confirmed plan can be undone"])

    drafts = ImportDraft.query.filter_by(plan_id=plan_id).all()
    po_imported = [d for d in drafts if d.file_type == DRAFT_FILE_TYPE_PO and d.status == DRAFT_STATUS_CONFIRMED]
    inv_imported = [d for d in drafts if d.file_type == DRAFT_FILE_TYPE_INVOICE and d.status == DRAFT_STATUS_CONFIRMED]

    po_before_ids = set()
    inv_before_ids = set()
    for snap in plan.snapshots:
        if not snap.restored:
            if snap.table_name == "purchase_orders":
                po_before_ids.add(snap.row_id)
            elif snap.table_name == "invoices":
                inv_before_ids.add(snap.row_id)

    if po_imported:
        all_po = PurchaseOrder.query.filter_by(batch_id=plan.batch_id).all()
        for rec in all_po:
            if rec.id not in po_before_ids:
                db.session.delete(rec)
        for snap in plan.snapshots:
            if snap.table_name == "purchase_orders" and not snap.restored and snap.original_data:
                orig = json.loads(snap.original_data)
                existing = PurchaseOrder.query.get(snap.row_id)
                if existing:
                    for k, v in orig.items():
                        if k not in ("id", "batch_id"):
                            setattr(existing, k, v)
                else:
                    db.session.add(PurchaseOrder(
                        id=snap.row_id,
                        batch_id=plan.batch_id,
                        po_number=orig.get("po_number", ""),
                        vendor_code=orig.get("vendor_code", ""),
                        vendor_name=orig.get("vendor_name", ""),
                        amount=orig.get("amount", 0),
                        currency=orig.get("currency", "CNY"),
                        po_date=orig.get("po_date", ""),
                        raw_data=orig.get("raw_data"),
                    ))
                snap.restored = True

    if inv_imported:
        all_inv = Invoice.query.filter_by(batch_id=plan.batch_id).all()
        for rec in all_inv:
            if rec.id not in inv_before_ids:
                db.session.delete(rec)
        for snap in plan.snapshots:
            if snap.table_name == "invoices" and not snap.restored and snap.original_data:
                orig = json.loads(snap.original_data)
                existing = Invoice.query.get(snap.row_id)
                if existing:
                    for k, v in orig.items():
                        if k not in ("id", "batch_id"):
                            setattr(existing, k, v)
                else:
                    db.session.add(Invoice(
                        id=snap.row_id,
                        batch_id=plan.batch_id,
                        invoice_number=orig.get("invoice_number", ""),
                        vendor_code=orig.get("vendor_code", ""),
                        vendor_name=orig.get("vendor_name", ""),
                        amount=orig.get("amount", 0),
                        currency=orig.get("currency", "CNY"),
                        invoice_date=orig.get("invoice_date", ""),
                        raw_data=orig.get("raw_data"),
                    ))
                snap.restored = True

    for d in drafts:
        if d.status == DRAFT_STATUS_CONFIRMED:
            d.status = DRAFT_STATUS_CANCELLED

    plan.status = PLAN_STATUS_UNDONE
    plan.undone_by = operator
    plan.undone_at = datetime.now(timezone.utc)

    db.session.add(AuditLog(
        batch_id=plan.batch_id,
        action="PLAN_UNDONE",
        detail=f"undo plan #{plan_id} by {operator}, data restored to pre-import state",
        operator=operator,
    ))
    db.session.commit()
    return {"success": True, "plan_id": plan_id, "note": "plan undone, data restored to pre-import state"}


def get_latest_plan_review_summary(batch_id):
    plan = ImportPlan.query.filter_by(
        batch_id=batch_id, status=PLAN_STATUS_CONFIRMED,
    ).order_by(ImportPlan.confirmed_at.desc()).first()
    if not plan or not plan.plan_summary:
        return None
    try:
        ps = json.loads(plan.plan_summary)
        parts = ps.get("parts", [])
        return f"plan #{plan.id}: {'; '.join(parts)} @{plan.confirmed_by or 'system'}"
    except Exception:
        return None


def get_health_rules(batch_id):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])
    rules = {r.rule_key: r for r in batch.health_check_rules}
    result = {}
    for key, default in DEFAULT_HEALTH_RULES.items():
        if key in rules:
            r = rules[key]
            result[key] = {
                "enabled": r.enabled,
                "severity": r.severity,
                "threshold": r.threshold,
                "rule_version": r.rule_version,
                "updated_by": r.updated_by,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
        else:
            result[key] = {
                "enabled": default["enabled"],
                "severity": default["severity"],
                "threshold": default["threshold"],
                "rule_version": None,
                "updated_by": "default",
                "updated_at": None,
            }
    def _normalize_threshold(v):
        try:
            fv = float(v)
            if fv.is_integer():
                return int(fv)
            return fv
        except (TypeError, ValueError):
            return v

    current_version = compute_health_rule_version(
        {k: {"e": v["enabled"], "s": v["severity"], "t": _normalize_threshold(v["threshold"])}
         for k, v in result.items()}
    )
    return {"rules": result, "rule_version": current_version}


def update_health_rules(batch_id, rules_update, operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])
    existing = {r.rule_key: r for r in batch.health_check_rules}
    for key, update in rules_update.items():
        if key not in DEFAULT_HEALTH_RULES:
            continue
        default = DEFAULT_HEALTH_RULES[key]
        enabled = bool(update.get("enabled", default["enabled"]))
        severity = update.get("severity", default["severity"])
        threshold = float(update.get("threshold", default["threshold"]))
        if severity not in (HEALTH_SEVERITY_BLOCKER, HEALTH_SEVERITY_WARNING, HEALTH_SEVERITY_INFO):
            severity = default["severity"]
        if key in existing:
            r = existing[key]
            r.enabled = enabled
            r.severity = severity
            r.threshold = threshold
            r.updated_by = operator
        else:
            r = HealthCheckRule(
                batch_id=batch_id,
                rule_key=key,
                enabled=enabled,
                severity=severity,
                threshold=threshold,
                rule_version="",
                updated_by=operator,
            )
            db.session.add(r)
    db.session.flush()
    all_rules = get_health_rules(batch_id)
    new_version = all_rules["rule_version"]
    for r in batch.health_check_rules:
        r.rule_version = new_version
    log = AuditLog(
        batch_id=batch_id,
        action="HEALTH_RULES_UPDATED",
        detail=f"health check rules updated by {operator}, version={new_version}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    return get_health_rules(batch_id)


def run_health_check(batch_id, operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])
    rules_info = get_health_rules(batch_id)
    rules = rules_info["rules"]
    rule_version = rules_info["rule_version"]
    results = []
    pos = batch.purchase_orders
    invs = batch.invoices
    source_files = []
    if batch.po_filename:
        source_files.append(batch.po_filename)
    if batch.invoice_filename:
        source_files.append(batch.invoice_filename)
    if rules.get(HEALTH_RULE_DUPLICATE_PO, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_DUPLICATE_PO]["severity"]
        po_num_map = {}
        for po in pos:
            if po.po_number not in po_num_map:
                po_num_map[po.po_number] = []
            po_num_map[po.po_number].append(po)
        for num, items in po_num_map.items():
            if len(items) > 1:
                results.append({
                    "rule_key": HEALTH_RULE_DUPLICATE_PO,
                    "severity": sev,
                    "category": "purchase_order",
                    "message": f"采购单号重复：{num}，共 {len(items)} 条",
                    "related_numbers": [num],
                    "table_name": "purchase_orders",
                    "row_id": items[0].id,
                })
    if rules.get(HEALTH_RULE_DUPLICATE_INVOICE, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_DUPLICATE_INVOICE]["severity"]
        inv_num_map = {}
        for inv in invs:
            if inv.invoice_number not in inv_num_map:
                inv_num_map[inv.invoice_number] = []
            inv_num_map[inv.invoice_number].append(inv)
        for num, items in inv_num_map.items():
            if len(items) > 1:
                results.append({
                    "rule_key": HEALTH_RULE_DUPLICATE_INVOICE,
                    "severity": sev,
                    "category": "invoice",
                    "message": f"发票号重复：{num}，共 {len(items)} 条",
                    "related_numbers": [num],
                    "table_name": "invoices",
                    "row_id": items[0].id,
                })
    if rules.get(HEALTH_RULE_MISSING_COLUMNS, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_MISSING_COLUMNS]["severity"]
        for po in pos:
            missing = []
            if not po.po_number:
                missing.append("po_number")
            if not po.vendor_code:
                missing.append("vendor_code")
            if not po.vendor_name:
                missing.append("vendor_name")
            if po.amount is None:
                missing.append("amount")
            if not po.po_date:
                missing.append("po_date")
            if missing:
                results.append({
                    "rule_key": HEALTH_RULE_MISSING_COLUMNS,
                    "severity": sev,
                    "category": "purchase_order",
                    "message": f"采购单 {po.po_number or '#'+str(po.id)} 缺少必填列：{', '.join(missing)}",
                    "related_numbers": [po.po_number] if po.po_number else [],
                    "table_name": "purchase_orders",
                    "row_id": po.id,
                })
        for inv in invs:
            missing = []
            if not inv.invoice_number:
                missing.append("invoice_number")
            if not inv.vendor_code:
                missing.append("vendor_code")
            if not inv.vendor_name:
                missing.append("vendor_name")
            if inv.amount is None:
                missing.append("amount")
            if not inv.invoice_date:
                missing.append("invoice_date")
            if missing:
                results.append({
                    "rule_key": HEALTH_RULE_MISSING_COLUMNS,
                    "severity": sev,
                    "category": "invoice",
                    "message": f"发票 {inv.invoice_number or '#'+str(inv.id)} 缺少必填列：{', '.join(missing)}",
                    "related_numbers": [inv.invoice_number] if inv.invoice_number else [],
                    "table_name": "invoices",
                    "row_id": inv.id,
                })
    if rules.get(HEALTH_RULE_NEGATIVE_AMOUNT, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_NEGATIVE_AMOUNT]["severity"]
        threshold = rules[HEALTH_RULE_NEGATIVE_AMOUNT]["threshold"]
        for po in pos:
            if po.amount is not None and po.amount < threshold:
                results.append({
                    "rule_key": HEALTH_RULE_NEGATIVE_AMOUNT,
                    "severity": sev,
                    "category": "purchase_order",
                    "message": f"采购单 {po.po_number} 金额 {po.amount} 低于阈值 {threshold}",
                    "related_numbers": [po.po_number],
                    "table_name": "purchase_orders",
                    "row_id": po.id,
                })
        for inv in invs:
            if inv.amount is not None and inv.amount < threshold:
                results.append({
                    "rule_key": HEALTH_RULE_NEGATIVE_AMOUNT,
                    "severity": sev,
                    "category": "invoice",
                    "message": f"发票 {inv.invoice_number} 金额 {inv.amount} 低于阈值 {threshold}",
                    "related_numbers": [inv.invoice_number],
                    "table_name": "invoices",
                    "row_id": inv.id,
                })
    if rules.get(HEALTH_RULE_VENDOR_MISMATCH, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_VENDOR_MISMATCH]["severity"]
        po_vendors = {}
        for po in pos:
            if po.vendor_code:
                if po.vendor_code not in po_vendors:
                    po_vendors[po.vendor_code] = set()
                po_vendors[po.vendor_code].add(po.vendor_name or "")
        for code, names in po_vendors.items():
            if len(names) > 1:
                results.append({
                    "rule_key": HEALTH_RULE_VENDOR_MISMATCH,
                    "severity": sev,
                    "category": "purchase_order",
                    "message": f"采购单供应商编码 {code} 对应多个名称：{', '.join(sorted(names))}",
                    "related_numbers": [code],
                    "table_name": "purchase_orders",
                    "row_id": None,
                })
        inv_vendors = {}
        for inv in invs:
            if inv.vendor_code:
                if inv.vendor_code not in inv_vendors:
                    inv_vendors[inv.vendor_code] = set()
                inv_vendors[inv.vendor_code].add(inv.vendor_name or "")
        for code, names in inv_vendors.items():
            if len(names) > 1:
                results.append({
                    "rule_key": HEALTH_RULE_VENDOR_MISMATCH,
                    "severity": sev,
                    "category": "invoice",
                    "message": f"发票供应商编码 {code} 对应多个名称：{', '.join(sorted(names))}",
                    "related_numbers": [code],
                    "table_name": "invoices",
                    "row_id": None,
                })
        po_codes = set(po.vendor_code for po in pos if po.vendor_code)
        inv_codes = set(inv.vendor_code for inv in invs if inv.vendor_code)
        only_po = po_codes - inv_codes
        only_inv = inv_codes - po_codes
        if only_po and invs:
            results.append({
                "rule_key": HEALTH_RULE_VENDOR_MISMATCH,
                "severity": sev,
                "category": "cross_table",
                "message": f"采购单有但发票无的供应商：{', '.join(sorted(only_po))}",
                "related_numbers": sorted(only_po),
                "table_name": None,
                "row_id": None,
            })
        if only_inv and pos:
            results.append({
                "rule_key": HEALTH_RULE_VENDOR_MISMATCH,
                "severity": sev,
                "category": "cross_table",
                "message": f"发票有但采购单无的供应商：{', '.join(sorted(only_inv))}",
                "related_numbers": sorted(only_inv),
                "table_name": None,
                "row_id": None,
            })
    if rules.get(HEALTH_RULE_CONFIRMED_OVERRIDE_RISK, {}).get("enabled", True):
        sev = rules[HEALTH_RULE_CONFIRMED_OVERRIDE_RISK]["severity"]
        other_batches = Batch.query.filter(
            Batch.id != batch_id,
            Batch.status == BATCH_STATUS_CONFIRMED,
        ).all()
        other_inv_nums = set()
        for ob in other_batches:
            for inv in ob.invoices:
                other_inv_nums.add(inv.invoice_number)
        current_inv_nums = set(inv.invoice_number for inv in invs)
        overlap = current_inv_nums & other_inv_nums
        if overlap:
            results.append({
                "rule_key": HEALTH_RULE_CONFIRMED_OVERRIDE_RISK,
                "severity": sev,
                "category": "cross_batch",
                "message": f"发现 {len(overlap)} 张发票已存在于其他已确认批次，存在覆盖风险：{', '.join(sorted(list(overlap))[:10])}{'...' if len(overlap) > 10 else ''}",
                "related_numbers": sorted(list(overlap))[:20],
                "table_name": "invoices",
                "row_id": None,
            })
    blocker_count = sum(1 for r in results if r["severity"] == HEALTH_SEVERITY_BLOCKER)
    warning_count = sum(1 for r in results if r["severity"] == HEALTH_SEVERITY_WARNING)
    info_count = sum(1 for r in results if r["severity"] == HEALTH_SEVERITY_INFO)
    summary = {
        "total": len(results),
        "blocker_count": blocker_count,
        "warning_count": warning_count,
        "info_count": info_count,
        "categories": {},
    }
    for r in results:
        cat = r["category"] or "other"
        if cat not in summary["categories"]:
            summary["categories"][cat] = {"total": 0, "blocker": 0, "warning": 0, "info": 0}
        summary["categories"][cat]["total"] += 1
        if r["severity"] == HEALTH_SEVERITY_BLOCKER:
            summary["categories"][cat]["blocker"] += 1
        elif r["severity"] == HEALTH_SEVERITY_WARNING:
            summary["categories"][cat]["warning"] += 1
        else:
            summary["categories"][cat]["info"] += 1
    history = HealthCheckHistory(
        batch_id=batch_id,
        rule_version=rule_version,
        operator=operator,
        source_files=json.dumps(source_files, ensure_ascii=False),
        summary=json.dumps(summary, ensure_ascii=False),
        blocker_count=blocker_count,
        warning_count=warning_count,
        info_count=info_count,
    )
    db.session.add(history)
    db.session.flush()
    for r in results:
        hr = HealthCheckResult(
            history_id=history.id,
            batch_id=batch_id,
            rule_key=r["rule_key"],
            severity=r["severity"],
            category=r["category"],
            message=r["message"],
            related_numbers=json.dumps(r["related_numbers"], ensure_ascii=False) if r["related_numbers"] else None,
            table_name=r["table_name"],
            row_id=r["row_id"],
        )
        db.session.add(hr)
    log = AuditLog(
        batch_id=batch_id,
        action="HEALTH_CHECK_RUN",
        detail=f"health check by {operator}, {blocker_count} blockers, {warning_count} warnings, {info_count} infos",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    return {
        "history_id": history.id,
        "rule_version": rule_version,
        "summary": summary,
        "results": [dict(r, id=None) for r in results],
        "created_at": history.created_at.isoformat() if history.created_at else None,
    }


def list_health_history(batch_id, limit=20):
    history = HealthCheckHistory.query.filter_by(batch_id=batch_id).order_by(
        HealthCheckHistory.created_at.desc()
    ).limit(limit).all()
    return [h.to_dict() for h in history]


def get_health_history_detail(history_id):
    history = HealthCheckHistory.query.get(history_id)
    if not history:
        raise ValidationError(["health check history not found"])
    results = HealthCheckResult.query.filter_by(history_id=history_id).order_by(
        HealthCheckResult.severity, HealthCheckResult.id
    ).all()
    data = history.to_dict()
    data["results"] = [r.to_dict() for r in results]
    return data


def export_health_check_csv(history_id):
    detail = get_health_history_detail(history_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["=== 数据健康巡检报告 ==="])
    writer.writerow(["巡检ID", detail["id"]])
    writer.writerow(["批次ID", detail["batch_id"]])
    writer.writerow(["规则版本", detail["rule_version"]])
    writer.writerow(["操作人", detail["operator"]])
    writer.writerow(["巡检时间", detail["created_at"]])
    writer.writerow(["阻断数", detail["blocker_count"]])
    writer.writerow(["警告数", detail["warning_count"]])
    writer.writerow(["提示数", detail["info_count"]])
    writer.writerow(["来源文件", ", ".join(detail.get("source_files", []))])
    writer.writerow([])
    writer.writerow(["=== 巡检问题明细 ==="])
    writer.writerow(["严重等级", "规则代码", "分类", "问题描述", "关联单号", "表名", "行ID"])
    for r in detail.get("results", []):
        writer.writerow([
            r["severity"],
            r["rule_key"],
            r.get("category", ""),
            r["message"],
            ", ".join(r.get("related_numbers", [])),
            r.get("table_name", ""),
            r.get("row_id", ""),
        ])
    writer.writerow([])
    writer.writerow(["=== 巡检摘要 ==="])
    summary = detail.get("summary", {})
    writer.writerow(["总问题数", summary.get("total", 0)])
    writer.writerow(["阻断", summary.get("blocker_count", 0)])
    writer.writerow(["警告", summary.get("warning_count", 0)])
    writer.writerow(["提示", summary.get("info_count", 0)])
    return output.getvalue()


def import_health_remarks(batch_id, csv_content, operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])
    rules_info = get_health_rules(batch_id)
    current_version = rules_info["rule_version"]
    reader = csv.reader(io.StringIO(csv_content))
    rows = list(reader)
    if not rows:
        raise ValidationError(["empty CSV file"])
    history_id = None
    csv_rule_version = None
    summary_found = False
    for i, row in enumerate(rows[:30]):
        if len(row) >= 2 and row[0] == "巡检ID":
            try:
                history_id = int(row[1])
            except (ValueError, IndexError):
                pass
        if len(row) >= 2 and row[0] == "规则版本":
            csv_rule_version = row[1].strip()
        if len(row) >= 1 and row[0] == "=== 巡检问题明细 ===":
            break
    if csv_rule_version and csv_rule_version != current_version:
        raise ValidationError([
            f"规则版本不一致：CSV 使用 {csv_rule_version}，当前批次为 {current_version}，请先同步规则后再导入"
        ])
    if history_id:
        existing = HealthCheckHistory.query.get(history_id)
        if existing and existing.batch_id != batch_id:
            raise ValidationError(["该巡检记录不属于当前批次，跨批次导入被阻断"])
        existing_remarks = NoteComparison.query.filter_by(
            batch_id=batch_id, change_source="HEALTH_REMARK_IMPORT"
        ).count()
        if existing_remarks > 0 and history_id:
            same_history = HealthCheckResult.query.filter_by(
                history_id=history_id, batch_id=batch_id
            ).first()
            if same_history:
                pass
    detail_start = -1
    for i, row in enumerate(rows):
        if len(row) >= 1 and row[0] == "=== 巡检问题明细 ===":
            detail_start = i
            break
    if detail_start < 0:
        raise ValidationError(["CSV 格式不正确，缺少巡检问题明细段"])
    issues = []
    header_idx = detail_start + 1
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if not row or (len(row) >= 1 and row[0].startswith("===")):
            break
        if len(row) < 4:
            continue
        severity = row[0].strip()
        rule_key = row[1].strip() if len(row) > 1 else ""
        category = row[2].strip() if len(row) > 2 else ""
        message = row[3].strip() if len(row) > 3 else ""
        if not message:
            continue
        issues.append({
            "severity": severity,
            "rule_key": rule_key,
            "category": category,
            "message": message,
        })
    if not issues:
        raise ValidationError(["CSV 中未找到有效的巡检问题"])
    remark_text = f"健康巡检导入备注（{len(issues)} 条问题）\n"
    for idx, issue in enumerate(issues[:50], 1):
        remark_text += f"{idx}. [{issue['severity']}] {issue['message']}\n"
    if len(issues) > 50:
        remark_text += f"... 共 {len(issues)} 条，仅显示前 50 条\n"
    log = AuditLog(
        batch_id=batch_id,
        action="HEALTH_REMARKS_IMPORTED",
        detail=f"imported {len(issues)} health check remarks from CSV by {operator}, rule_version={csv_rule_version or 'unknown'}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    return {
        "imported": len(issues),
        "summary": remark_text,
        "rule_version_match": csv_rule_version == current_version if csv_rule_version else None,
        "history_id": history_id,
    }


def _generate_handover_list_number(batch_id):
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y%m%d")
    prefix = f"HO-{batch_id:04d}-{date_str}-"
    latest = HandoverList.query.filter(
        HandoverList.list_number.like(prefix + "%")
    ).order_by(HandoverList.id.desc()).first()
    seq = 1
    if latest:
        try:
            seq = int(latest.list_number.split("-")[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f"{prefix}{seq:04d}"


def _get_latest_confirmed_plan(batch_id):
    from models import PLAN_STATUS_CONFIRMED
    plan = ImportPlan.query.filter_by(
        batch_id=batch_id, status=PLAN_STATUS_CONFIRMED,
    ).order_by(ImportPlan.confirmed_at.desc()).first()
    if plan:
        ps = json.loads(plan.plan_summary) if plan.plan_summary else {}
        return {
            "plan_id": plan.id,
            "status": plan.status,
            "confirmed_by": plan.confirmed_by,
            "confirmed_at": plan.confirmed_at.isoformat() if plan.confirmed_at else None,
            "summary": ps,
        }
    return None


def _get_latest_health_summary(batch_id):
    history = HealthCheckHistory.query.filter_by(
        batch_id=batch_id,
    ).order_by(HealthCheckHistory.created_at.desc()).first()
    if history:
        return {
            "history_id": history.id,
            "rule_version": history.rule_version,
            "operator": history.operator,
            "blocker_count": history.blocker_count,
            "warning_count": history.warning_count,
            "info_count": history.info_count,
            "summary": json.loads(history.summary) if history.summary else None,
            "created_at": history.created_at.isoformat() if history.created_at else None,
        }, history.id
    return None, None


def create_handover_list(batch_id, pending_remarks="", operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])

    content_hash = compute_handover_content_hash(batch)

    list_number = _generate_handover_list_number(batch_id)

    summary = batch.summary

    latest_plan = _get_latest_confirmed_plan(batch_id)
    latest_health, latest_health_id = _get_latest_health_summary(batch_id)

    handover = HandoverList(
        batch_id=batch_id,
        list_number=list_number,
        status=HANDOVER_STATUS_DRAFT,
        batch_status=batch.status,
        payable_total=summary["payable_total"],
        matched_count=summary["matched_count"],
        exception_count=summary["exception_count"],
        unmatched_po_count=summary["unmatched_po_count"],
        unmatched_invoice_count=summary["unmatched_invoice_count"],
        latest_import_plan=json.dumps(latest_plan, ensure_ascii=False) if latest_plan else None,
        latest_health_summary=json.dumps(latest_health, ensure_ascii=False) if latest_health else None,
        latest_health_history_id=latest_health_id,
        export_filename=None,
        pending_remarks=pending_remarks,
        batch_updated_at=batch.updated_at,
        content_hash=content_hash,
        created_by=operator,
    )
    db.session.add(handover)
    db.session.flush()

    results = MatchResult.query.filter_by(batch_id=batch_id).order_by(MatchResult.id).all()
    for idx, mr in enumerate(results):
        exc_remark = ""
        for e in batch.exception_items:
            if e.match_result_id == mr.id and e.remarks:
                exc_remark = e.remarks
                break
        item = HandoverListItem(
            handover_list_id=handover.id,
            match_result_id=mr.id,
            po_number=mr.po.po_number if mr.po else None,
            invoice_number=mr.invoice.invoice_number if mr.invoice else None,
            vendor_code=mr.po.vendor_code if mr.po else (mr.invoice.vendor_code if mr.invoice else None),
            vendor_name=mr.po.vendor_name if mr.po else (mr.invoice.vendor_name if mr.invoice else None),
            po_amount=mr.po_amount,
            invoice_amount=mr.invoice_amount,
            amount_diff=mr.amount_diff,
            match_type=mr.match_type,
            is_exception=mr.is_exception,
            exception_type=mr.exception_type,
            status=mr.status,
            remarks=exc_remark if exc_remark else (mr.remarks or ""),
            rule_version=mr.rule_version,
            item_order=idx,
        )
        db.session.add(item)

    log = AuditLog(
        batch_id=batch_id,
        action="HANDOVER_CREATE",
        detail=f"创建对账交接清单 #{handover.id} ({list_number})，共 {len(results)} 条明细",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return handover


def get_handover_list(handover_id):
    return HandoverList.query.get(handover_id)


def get_handover_list_by_number(list_number):
    return HandoverList.query.filter_by(list_number=list_number).first()


def list_handover_lists(batch_id=None, status=None):
    query = HandoverList.query
    if batch_id is not None:
        query = query.filter_by(batch_id=batch_id)
    if status:
        query = query.filter_by(status=status)
    return query.order_by(HandoverList.created_at.desc()).all()


def get_handover_items(handover_id):
    return HandoverListItem.query.filter_by(
        handover_list_id=handover_id
    ).order_by(HandoverListItem.item_order).all()


def refresh_handover_list(handover_id, operator="system"):
    handover = HandoverList.query.get(handover_id)
    if not handover:
        raise ValidationError(["handover list not found"])
    if handover.status != HANDOVER_STATUS_DRAFT:
        raise ValidationError([f"只有草稿状态的清单可以刷新，当前状态: {handover.status}"])

    batch = Batch.query.get(handover.batch_id)
    if not batch:
        raise ValidationError(["batch not found"])

    content_hash = compute_handover_content_hash(batch)

    if content_hash == handover.content_hash:
        return handover, False

    summary = batch.summary
    latest_plan = _get_latest_confirmed_plan(handover.batch_id)
    latest_health, latest_health_id = _get_latest_health_summary(handover.batch_id)

    handover.batch_status = batch.status
    handover.payable_total = summary["payable_total"]
    handover.matched_count = summary["matched_count"]
    handover.exception_count = summary["exception_count"]
    handover.unmatched_po_count = summary["unmatched_po_count"]
    handover.unmatched_invoice_count = summary["unmatched_invoice_count"]
    handover.latest_import_plan = json.dumps(latest_plan, ensure_ascii=False) if latest_plan else None
    handover.latest_health_summary = json.dumps(latest_health, ensure_ascii=False) if latest_health else None
    handover.latest_health_history_id = latest_health_id
    handover.batch_updated_at = batch.updated_at
    handover.content_hash = content_hash

    HandoverListItem.query.filter_by(handover_list_id=handover_id).delete()

    results = MatchResult.query.filter_by(batch_id=handover.batch_id).order_by(MatchResult.id).all()
    for idx, mr in enumerate(results):
        exc_remark = ""
        for e in batch.exception_items:
            if e.match_result_id == mr.id and e.remarks:
                exc_remark = e.remarks
                break
        item = HandoverListItem(
            handover_list_id=handover.id,
            match_result_id=mr.id,
            po_number=mr.po.po_number if mr.po else None,
            invoice_number=mr.invoice.invoice_number if mr.invoice else None,
            vendor_code=mr.po.vendor_code if mr.po else (mr.invoice.vendor_code if mr.invoice else None),
            vendor_name=mr.po.vendor_name if mr.po else (mr.invoice.vendor_name if mr.invoice else None),
            po_amount=mr.po_amount,
            invoice_amount=mr.invoice_amount,
            amount_diff=mr.amount_diff,
            match_type=mr.match_type,
            is_exception=mr.is_exception,
            exception_type=mr.exception_type,
            status=mr.status,
            remarks=exc_remark if exc_remark else (mr.remarks or ""),
            rule_version=mr.rule_version,
            item_order=idx,
        )
        db.session.add(item)

    log = AuditLog(
        batch_id=handover.batch_id,
        action="HANDOVER_REFRESH",
        detail=f"刷新对账交接清单 #{handover.id} ({handover.list_number})，更新为 {len(results)} 条明细",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return handover, True


def complete_handover_list(handover_id, role="viewer", operator="system"):
    handover = HandoverList.query.get(handover_id)
    if not handover:
        raise ValidationError(["handover list not found"])

    if role not in HANDOVER_PERMISSION_COMPLETE:
        log = AuditLog(
            batch_id=handover.batch_id,
            action="HANDOVER_COMPLETE_DENIED",
            detail=f"用户 {operator} (角色 {role}) 无权限完成对账交接清单 #{handover.id}",
            operator=operator,
        )
        db.session.add(log)
        db.session.commit()
        raise ValidationError([f"权限不足：只有 {HANDOVER_PERMISSION_COMPLETE} 角色可以完成清单，当前角色: {role}"])

    if not handover.can_transition(HANDOVER_STATUS_COMPLETED):
        raise ValidationError([f"清单状态 '{handover.status}' 不允许完成操作"])

    current_hash = compute_handover_content_hash(Batch.query.get(handover.batch_id))
    if current_hash != handover.content_hash:
        log = AuditLog(
            batch_id=handover.batch_id,
            action="HANDOVER_COMPLETE_DENIED",
            detail=f"批次已变更，清单 #{handover.id} 未刷新，完成操作被拒绝",
            operator=operator,
        )
        db.session.add(log)
        db.session.commit()
        raise ValidationError(["批次数据已变更，请先刷新清单后再完成"])

    handover.status = HANDOVER_STATUS_COMPLETED
    handover.completed_by = operator
    handover.completed_at = datetime.now(timezone.utc)

    log = AuditLog(
        batch_id=handover.batch_id,
        action="HANDOVER_COMPLETE",
        detail=f"完成对账交接清单 #{handover.id} ({handover.list_number})",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return handover


def void_handover_list(handover_id, reason="", role="viewer", operator="system"):
    handover = HandoverList.query.get(handover_id)
    if not handover:
        raise ValidationError(["handover list not found"])

    if role not in HANDOVER_PERMISSION_VOID:
        log = AuditLog(
            batch_id=handover.batch_id,
            action="HANDOVER_VOID_DENIED",
            detail=f"用户 {operator} (角色 {role}) 无权限作废对账交接清单 #{handover.id}",
            operator=operator,
        )
        db.session.add(log)
        db.session.commit()
        raise ValidationError([f"权限不足：只有 {HANDOVER_PERMISSION_VOID} 角色可以作废清单，当前角色: {role}"])

    if not handover.can_transition(HANDOVER_STATUS_VOID):
        raise ValidationError([f"清单状态 '{handover.status}' 不允许作废操作"])

    handover.status = HANDOVER_STATUS_VOID
    handover.voided_by = operator
    handover.voided_at = datetime.now(timezone.utc)
    handover.void_reason = reason

    log = AuditLog(
        batch_id=handover.batch_id,
        action="HANDOVER_VOID",
        detail=f"作废对账交接清单 #{handover.id} ({handover.list_number})，原因: {reason}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return handover


def export_handover_csv(handover_id):
    handover = HandoverList.query.get(handover_id)
    if not handover:
        raise ValidationError(["handover list not found"])

    items = get_handover_items(handover_id)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["=== 对账交接清单头 ==="])
    writer.writerow(["清单编号", handover.list_number])
    writer.writerow(["清单状态", handover.status])
    writer.writerow(["批次ID", handover.batch_id])
    writer.writerow(["批次状态", handover.batch_status])
    writer.writerow(["应付合计", handover.payable_total])
    writer.writerow(["匹配成功数", handover.matched_count])
    writer.writerow(["异常数量", handover.exception_count])
    writer.writerow(["未匹配采购单数", handover.unmatched_po_count])
    writer.writerow(["未匹配发票数", handover.unmatched_invoice_count])

    if handover.latest_import_plan:
        plan = json.loads(handover.latest_import_plan)
        writer.writerow(["最近导入方案ID", plan.get("plan_id", "")])
        writer.writerow(["最近导入方案确认人", plan.get("confirmed_by", "")])
        writer.writerow(["最近导入方案确认时间", plan.get("confirmed_at", "")])

    if handover.latest_health_summary:
        health = json.loads(handover.latest_health_summary)
        writer.writerow(["最近巡检ID", health.get("history_id", "")])
        writer.writerow(["最近巡检摘要", f"阻断:{health.get('blocker_count',0)} 警告:{health.get('warning_count',0)} 信息:{health.get('info_count',0)}"])

    writer.writerow(["导出文件名", handover.export_filename or ""])
    writer.writerow(["待处理备注", handover.pending_remarks or ""])
    writer.writerow(["批次更新时间", handover.batch_updated_at.isoformat() if handover.batch_updated_at else ""])
    writer.writerow(["内容哈希", handover.content_hash])
    writer.writerow(["创建人", handover.created_by])
    writer.writerow(["创建时间", handover.created_at.isoformat() if handover.created_at else ""])
    writer.writerow(["完成人", handover.completed_by or ""])
    writer.writerow(["完成时间", handover.completed_at.isoformat() if handover.completed_at else ""])
    writer.writerow(["作废人", handover.voided_by or ""])
    writer.writerow(["作废时间", handover.voided_at.isoformat() if handover.voided_at else ""])
    writer.writerow(["作废原因", handover.void_reason or ""])
    writer.writerow([])

    writer.writerow(["=== 对账交接清单明细 ==="])
    writer.writerow([
        "序号", "匹配结果ID", "采购单号", "发票号", "供应商编码", "供应商名称",
        "采购金额", "发票金额", "金额差异", "匹配类型", "是否异常",
        "异常类型", "匹配状态", "规则版本", "备注",
    ])
    for idx, item in enumerate(items, 1):
        writer.writerow([
            idx,
            item.match_result_id or "",
            item.po_number or "",
            item.invoice_number or "",
            item.vendor_code or "",
            item.vendor_name or "",
            f"{item.po_amount:.2f}" if item.po_amount is not None else "",
            f"{item.invoice_amount:.2f}" if item.invoice_amount is not None else "",
            f"{item.amount_diff:.2f}" if item.amount_diff is not None else "",
            item.match_type or "",
            "是" if item.is_exception else "否",
            item.exception_type or "",
            item.status or "",
            item.rule_version or "",
            item.remarks or "",
        ])

    output.seek(0)
    return output.getvalue()


def import_handover_csv(batch_id, csv_content, operator="system"):
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["batch not found"])

    reader = csv.reader(io.StringIO(csv_content))
    rows = list(reader)
    if not rows:
        raise ValidationError(["empty CSV file"])

    header_data = {}
    detail_start = -1
    list_number = None

    for i, row in enumerate(rows):
        if len(row) >= 1 and row[0] == "=== 对账交接清单头 ===":
            continue
        if len(row) >= 1 and row[0] == "=== 对账交接清单明细 ===":
            detail_start = i
            break
        if len(row) >= 2 and not row[0].startswith("===") and row[0] != "":
            key = row[0].strip()
            val = row[1].strip() if len(row) > 1 else ""
            header_data[key] = val
            if key == "清单编号":
                list_number = val

    if not list_number:
        raise ValidationError(["CSV 格式错误，缺少清单编号"])

    existing = get_handover_list_by_number(list_number)
    if existing:
        log = AuditLog(
            batch_id=batch_id,
            action="HANDOVER_IMPORT_DENIED",
            detail=f"重复回导被拒绝：清单编号 {list_number} 已存在（#{existing.id}）",
            operator=operator,
        )
        db.session.add(log)
        db.session.commit()
        raise ValidationError([f"清单编号 {list_number} 已存在，禁止重复回导"])

    content_hash = header_data.get("内容哈希", "")
    current_hash = compute_handover_content_hash(batch)

    if content_hash != current_hash:
        log = AuditLog(
            batch_id=batch_id,
            action="HANDOVER_IMPORT_DENIED",
            detail=f"批次已变更，回导被拒绝：清单 {list_number} 哈希不匹配",
            operator=operator,
        )
        db.session.add(log)
        db.session.commit()
        raise ValidationError(["批次数据已变更，与清单快照不一致，禁止回导。请先确认批次数据与清单导出时一致。"])

    if detail_start < 0:
        raise ValidationError(["CSV 格式错误，缺少明细段"])

    detail_items = []
    for i in range(detail_start + 2, len(rows)):
        row = rows[i]
        if not row or (len(row) >= 1 and row[0].startswith("===")):
            break
        if len(row) < 2:
            continue
        detail_items.append(row)

    handover = HandoverList(
        batch_id=batch_id,
        list_number=list_number,
        status=HANDOVER_STATUS_DRAFT,
        batch_status=header_data.get("批次状态", ""),
        payable_total=float(header_data.get("应付合计", 0)) if header_data.get("应付合计", "") else 0.0,
        matched_count=int(header_data.get("匹配成功数", 0)) if header_data.get("匹配成功数", "") else 0,
        exception_count=int(header_data.get("异常数量", 0)) if header_data.get("异常数量", "") else 0,
        unmatched_po_count=int(header_data.get("未匹配采购单数", 0)) if header_data.get("未匹配采购单数", "") else 0,
        unmatched_invoice_count=int(header_data.get("未匹配发票数", 0)) if header_data.get("未匹配发票数", "") else 0,
        export_filename=header_data.get("导出文件名", "") or None,
        pending_remarks=header_data.get("待处理备注", "") or None,
        content_hash=content_hash,
        created_by=header_data.get("创建人", operator) or operator,
    )
    db.session.add(handover)
    db.session.flush()

    for idx, row in enumerate(detail_items):
        def safe_get(arr, i, default=""):
            return arr[i].strip() if i < len(arr) else default

        is_exception = safe_get(row, 10) == "是"
        po_amount = safe_get(row, 6)
        inv_amount = safe_get(row, 7)
        amt_diff = safe_get(row, 8)

        item = HandoverListItem(
            handover_list_id=handover.id,
            match_result_id=int(safe_get(row, 1)) if safe_get(row, 1).isdigit() else None,
            po_number=safe_get(row, 2) or None,
            invoice_number=safe_get(row, 3) or None,
            vendor_code=safe_get(row, 4) or None,
            vendor_name=safe_get(row, 5) or None,
            po_amount=float(po_amount) if po_amount else None,
            invoice_amount=float(inv_amount) if inv_amount else None,
            amount_diff=float(amt_diff) if amt_diff else None,
            match_type=safe_get(row, 9) or None,
            is_exception=is_exception,
            exception_type=safe_get(row, 11) or None,
            status=safe_get(row, 12) or None,
            rule_version=safe_get(row, 13) or None,
            remarks=safe_get(row, 14) or None,
            item_order=idx,
        )
        db.session.add(item)

    log = AuditLog(
        batch_id=batch_id,
        action="HANDOVER_IMPORT",
        detail=f"回导对账交接清单 {list_number} (#{handover.id})，共 {len(detail_items)} 条明细",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return handover


def list_handover_audit_logs(batch_id=None, handover_id=None, limit=50):
    query = AuditLog.query
    if batch_id is not None:
        query = query.filter_by(batch_id=batch_id)
    action_filter = "HANDOVER_"
    query = query.filter(AuditLog.action.like(action_filter + "%"))
    if handover_id is not None:
        query = query.filter(AuditLog.detail.like(f"%#{handover_id}%"))
    return query.order_by(AuditLog.created_at.desc()).limit(limit).all()


def _generate_release_package_number(batch_id):
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = ReleasePackage.query.filter(
        ReleasePackage.batch_id == batch_id,
        ReleasePackage.package_number.like(f"REL-{batch_id}-{today}%"),
    ).count()
    return f"REL-{batch_id}-{today}-{count + 1:03d}"


def check_release_expiry(batch_id, operator="system"):
    """检查批次所有待审批放行包是否过期，过期则标记并写审计日志"""
    batch = Batch.query.get(batch_id)
    if not batch:
        return []

    pending_packages = ReleasePackage.query.filter_by(
        batch_id=batch_id,
        status=RELEASE_STATUS_PENDING,
        is_expired=False,
    ).all()

    expired = []
    latest_note = get_latest_recalc_note(batch_id)
    latest_health = (
        HealthCheckHistory.query.filter_by(batch_id=batch_id)
        .order_by(HealthCheckHistory.created_at.desc())
        .first()
    )

    for pkg in pending_packages:
        reasons = []
        if pkg.rule_version != batch.rule_version:
            reasons.append(f"容差规则变更: {pkg.rule_version[:8]} → {batch.rule_version[:8]}")
        if pkg.batch_status != batch.status:
            reasons.append(f"批次状态变更: {pkg.batch_status} → {batch.status}")
        if latest_note and pkg.recalc_note_id != latest_note.id:
            reasons.append(f"重算说明变更: v{pkg.recalc_note_version} → v{latest_note.version}")
        if latest_health and pkg.health_history_id != latest_health.id:
            reasons.append(f"巡检结果变更: 巡检#{pkg.health_history_id} → #{latest_health.id}")

        current_hash = compute_release_content_hash(batch, latest_note, latest_health)
        if pkg.content_hash != current_hash:
            if not reasons:
                reasons.append("批次数据内容发生变化")

        if reasons:
            pkg.is_expired = True
            pkg.status = RELEASE_STATUS_EXPIRED
            pkg.expire_reason = "; ".join(reasons)
            pkg.expired_at = datetime.now(timezone.utc)
            expired.append(pkg)

            log = AuditLog(
                batch_id=batch_id,
                action="RELEASE_EXPIRED",
                detail=f"放行包 {pkg.package_number} (#{pkg.id}) 已过期: {'; '.join(reasons)}",
                operator=operator,
            )
            db.session.add(log)

    if expired:
        db.session.commit()

    return expired


def mark_expired_packages_on_change(batch_id, change_source, operator="system"):
    """在批次数据、容差、巡检、重算说明变化时调用，标记旧包过期"""
    return check_release_expiry(batch_id, operator=operator)


def create_release_package(batch_id, remarks="", operator="system"):
    """创建放行包，自动快照当前批次状态"""
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    if batch.status not in (BATCH_STATUS_CONFIRMED,):
        raise ValidationError([f"批次状态 '{batch.status}' 不允许创建放行包，需先确认批次"])

    check_release_expiry(batch_id, operator=operator)

    summary = batch.summary
    latest_note = get_latest_recalc_note(batch_id)
    latest_health = (
        HealthCheckHistory.query.filter_by(batch_id=batch_id)
        .order_by(HealthCheckHistory.created_at.desc())
        .first()
    )
    latest_plan = (
        ImportPlan.query.filter_by(batch_id=batch_id, status=PLAN_STATUS_CONFIRMED)
        .order_by(ImportPlan.confirmed_at.desc())
        .first()
    )

    content_hash = compute_release_content_hash(batch, latest_note, latest_health)

    existing = ReleasePackage.query.filter_by(
        batch_id=batch_id,
        content_hash=content_hash,
        is_expired=False,
    ).first()
    if existing:
        return existing, False

    package_number = _generate_release_package_number(batch_id)

    pkg = ReleasePackage(
        batch_id=batch_id,
        package_number=package_number,
        status=RELEASE_STATUS_DRAFT,
        batch_status=batch.status,
        payable_total=summary["payable_total"],
        matched_count=summary["matched_count"],
        exception_count=summary["exception_count"],
        unmatched_po_count=summary["unmatched_po_count"],
        unmatched_invoice_count=summary["unmatched_invoice_count"],
        tolerance_pct=batch.tolerance_pct,
        tolerance_abs=batch.tolerance_abs,
        rule_version=batch.rule_version,
        recalc_note_id=latest_note.id if latest_note else None,
        recalc_note_version=latest_note.version if latest_note else None,
        recalc_note_summary=latest_note.change_summary if latest_note else None,
        health_history_id=latest_health.id if latest_health else None,
        health_rule_version=latest_health.rule_version if latest_health else None,
        health_summary=json.dumps(latest_health.summary, ensure_ascii=False) if latest_health and latest_health.summary else None,
        health_blocker_count=latest_health.blocker_count if latest_health else 0,
        health_warning_count=latest_health.warning_count if latest_health else 0,
        health_info_count=latest_health.info_count if latest_health else 0,
        import_plan_id=latest_plan.id if latest_plan else None,
        import_plan_summary=latest_plan.plan_summary if latest_plan else None,
        content_hash=content_hash,
        remarks=remarks,
        created_by=operator,
    )
    db.session.add(pkg)
    db.session.flush()

    matched_types = (MATCH_TYPE_EXACT, MATCH_TYPE_TOLERANCE, MATCH_TYPE_OVER_TOLERANCE)
    matched_results = [r for r in batch.match_results if r.match_type in matched_types]

    exc_map = {}
    for exc in batch.exception_items:
        if exc.match_result_id:
            exc_map[exc.match_result_id] = exc

    for idx, mr in enumerate(matched_results):
        exc = exc_map.get(mr.id)
        item = ReleasePackageItem(
            release_package_id=pkg.id,
            match_result_id=mr.id,
            po_number=mr.po.po_number if mr.po else None,
            invoice_number=mr.invoice.invoice_number if mr.invoice else None,
            vendor_code=mr.po.vendor_code if mr.po else (mr.invoice.vendor_code if mr.invoice else None),
            vendor_name=mr.po.vendor_name if mr.po else (mr.invoice.vendor_name if mr.invoice else None),
            po_amount=mr.po_amount,
            invoice_amount=mr.invoice_amount,
            amount_diff=mr.amount_diff,
            match_type=mr.match_type,
            is_exception=mr.is_exception,
            exception_type=mr.exception_type,
            status=mr.status,
            remarks=mr.remarks,
            exception_remarks=exc.remarks if exc else None,
            rule_version=mr.rule_version,
            item_order=idx,
        )
        db.session.add(item)

    log = AuditLog(
        batch_id=batch_id,
        action="RELEASE_CREATE",
        detail=f"创建放行包 {package_number} (#{pkg.id})，应付合计 {summary['payable_total']:.2f}，共 {len(matched_results)} 条明细",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    db.session.refresh(pkg)

    return pkg, True


def get_release_package(package_id):
    return ReleasePackage.query.get(package_id)


def get_release_package_by_number(package_number):
    return ReleasePackage.query.filter_by(package_number=package_number).first()


def list_release_packages(batch_id=None, status=None):
    query = ReleasePackage.query
    if batch_id is not None:
        query = query.filter_by(batch_id=batch_id)
    if status:
        query = query.filter_by(status=status)
    return query.order_by(ReleasePackage.created_at.desc()).all()


def get_latest_release_package(batch_id):
    return (
        ReleasePackage.query.filter_by(batch_id=batch_id)
        .order_by(ReleasePackage.created_at.desc())
        .first()
    )


def get_release_package_items(package_id):
    return ReleasePackageItem.query.filter_by(
        release_package_id=package_id
    ).order_by(ReleasePackageItem.item_order).all()


def submit_release_package(package_id, operator="system"):
    """提交放行包审批：DRAFT → PENDING"""
    pkg = get_release_package(package_id)
    if not pkg:
        raise ValidationError(["放行包不存在"])

    if pkg.is_expired:
        raise ValidationError([f"放行包已过期: {pkg.expire_reason}"])

    if pkg.status != RELEASE_STATUS_DRAFT:
        raise ValidationError([f"放行包状态 '{pkg.status}' 不允许提交，需为 DRAFT"])

    check_release_expiry(pkg.batch_id, operator=operator)
    pkg = get_release_package(package_id)
    if pkg.is_expired:
        raise ValidationError([f"放行包已过期: {pkg.expire_reason}"])

    pkg.status = RELEASE_STATUS_PENDING
    pkg.submitted_by = operator
    pkg.submitted_at = datetime.now(timezone.utc)

    log = AuditLog(
        batch_id=pkg.batch_id,
        action="RELEASE_SUBMIT",
        detail=f"提交放行包 {pkg.package_number} (#{pkg.id}) 审批",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return pkg


def approve_release_package(package_id, role="viewer", operator="system"):
    """审批通过放行包"""
    pkg = get_release_package(package_id)
    if not pkg:
        raise ValidationError(["放行包不存在"])

    if role not in RELEASE_PERMISSION_APPROVE:
        raise ValidationError([f"角色 '{role}' 无审批权限"])

    if pkg.is_expired:
        raise ValidationError([f"放行包已过期: {pkg.expire_reason}"])

    check_release_expiry(pkg.batch_id, operator=operator)
    pkg = get_release_package(package_id)
    if pkg.is_expired:
        raise ValidationError([f"放行包已过期，无法审批: {pkg.expire_reason}"])

    if not pkg.can_transition(RELEASE_STATUS_APPROVED):
        raise ValidationError([f"放行包状态 '{pkg.status}' 不允许审批通过"])

    pkg.status = RELEASE_STATUS_APPROVED
    pkg.approved_by = operator
    pkg.approved_at = datetime.now(timezone.utc)

    log = AuditLog(
        batch_id=pkg.batch_id,
        action="RELEASE_APPROVE",
        detail=f"审批通过放行包 {pkg.package_number} (#{pkg.id})",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return pkg


def reject_release_package(package_id, reason="", role="viewer", operator="system"):
    """审批驳回放行包"""
    pkg = get_release_package(package_id)
    if not pkg:
        raise ValidationError(["放行包不存在"])

    if role not in RELEASE_PERMISSION_REJECT:
        raise ValidationError([f"角色 '{role}' 无驳回权限"])

    if pkg.is_expired:
        raise ValidationError([f"放行包已过期: {pkg.expire_reason}"])

    if not pkg.can_transition(RELEASE_STATUS_REJECTED):
        raise ValidationError([f"放行包状态 '{pkg.status}' 不允许驳回"])

    pkg.status = RELEASE_STATUS_REJECTED
    pkg.rejected_by = operator
    pkg.rejected_at = datetime.now(timezone.utc)
    pkg.reject_reason = reason

    log = AuditLog(
        batch_id=pkg.batch_id,
        action="RELEASE_REJECT",
        detail=f"驳回放行包 {pkg.package_number} (#{pkg.id}): {reason}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return pkg


def revoke_release_package(package_id, reason="", role="viewer", operator="system"):
    """撤销放行包"""
    pkg = get_release_package(package_id)
    if not pkg:
        raise ValidationError(["放行包不存在"])

    if role not in RELEASE_PERMISSION_REVOKE:
        raise ValidationError([f"角色 '{role}' 无撤销权限"])

    if not pkg.can_transition(RELEASE_STATUS_REVOKED):
        raise ValidationError([f"放行包状态 '{pkg.status}' 不允许撤销"])

    pkg.status = RELEASE_STATUS_REVOKED
    pkg.revoked_by = operator
    pkg.revoked_at = datetime.now(timezone.utc)
    pkg.revoke_reason = reason

    log = AuditLog(
        batch_id=pkg.batch_id,
        action="RELEASE_REVOKE",
        detail=f"撤销放行包 {pkg.package_number} (#{pkg.id}): {reason}",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()

    return pkg


def export_release_csv(package_id):
    """导出行包CSV，带头段和明细"""
    pkg = get_release_package(package_id)
    if not pkg:
        raise ValidationError(["放行包不存在"])

    items = get_release_package_items(package_id)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["===== 放行包头段 ====="])
    writer.writerow(["放行包编号", pkg.package_number])
    writer.writerow(["放行包ID", pkg.id])
    writer.writerow(["批次ID", pkg.batch_id])
    writer.writerow(["状态", pkg.status])
    writer.writerow(["是否过期", "是" if pkg.is_expired else "否"])
    if pkg.is_expired:
        writer.writerow(["过期原因", pkg.expire_reason or ""])
    writer.writerow(["批次状态快照", pkg.batch_status or ""])
    writer.writerow(["应付合计", f"{pkg.payable_total:.2f}" if pkg.payable_total is not None else ""])
    writer.writerow(["匹配成功数", pkg.matched_count or 0])
    writer.writerow(["异常数", pkg.exception_count or 0])
    writer.writerow(["未匹配采购单数", pkg.unmatched_po_count or 0])
    writer.writerow(["未匹配发票数", pkg.unmatched_invoice_count or 0])
    writer.writerow(["容差百分比", pkg.tolerance_pct or ""])
    writer.writerow(["容差绝对值", pkg.tolerance_abs or ""])
    writer.writerow(["规则版本", pkg.rule_version or ""])
    writer.writerow(["重算说明ID", pkg.recalc_note_id or ""])
    writer.writerow(["重算说明版本", pkg.recalc_note_version or ""])
    writer.writerow(["重算说明摘要", pkg.recalc_note_summary or ""])
    writer.writerow(["巡检记录ID", pkg.health_history_id or ""])
    writer.writerow(["巡检规则版本", pkg.health_rule_version or ""])
    writer.writerow(["巡检阻断数", pkg.health_blocker_count or 0])
    writer.writerow(["巡检警告数", pkg.health_warning_count or 0])
    writer.writerow(["巡检提示数", pkg.health_info_count or 0])
    writer.writerow(["导入方案ID", pkg.import_plan_id or ""])
    writer.writerow(["经办人", pkg.created_by or ""])
    writer.writerow(["提交人", pkg.submitted_by or ""])
    writer.writerow(["提交时间", pkg.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if pkg.submitted_at else ""])
    writer.writerow(["审批人", pkg.approved_by or ""])
    writer.writerow(["审批时间", pkg.approved_at.strftime("%Y-%m-%d %H:%M:%S") if pkg.approved_at else ""])
    writer.writerow(["驳回人", pkg.rejected_by or ""])
    writer.writerow(["驳回时间", pkg.rejected_at.strftime("%Y-%m-%d %H:%M:%S") if pkg.rejected_at else ""])
    writer.writerow(["驳回原因", pkg.reject_reason or ""])
    writer.writerow(["撤销人", pkg.revoked_by or ""])
    writer.writerow(["撤销时间", pkg.revoked_at.strftime("%Y-%m-%d %H:%M:%S") if pkg.revoked_at else ""])
    writer.writerow(["撤销原因", pkg.revoke_reason or ""])
    writer.writerow(["内容哈希", pkg.content_hash or ""])
    writer.writerow(["备注", pkg.remarks or ""])
    writer.writerow(["创建时间", pkg.created_at.strftime("%Y-%m-%d %H:%M:%S") if pkg.created_at else ""])

    writer.writerow([])
    writer.writerow(["===== 放行包明细 ====="])
    writer.writerow([
        "序号", "匹配结果ID", "采购单号", "发票号", "供应商编码", "供应商名称",
        "采购金额", "发票金额", "金额差异", "匹配类型", "是否异常", "异常类型",
        "匹配状态", "规则版本", "匹配备注", "异常备注",
    ])

    for item in items:
        writer.writerow([
            item.item_order + 1,
            item.match_result_id or "",
            item.po_number or "",
            item.invoice_number or "",
            item.vendor_code or "",
            item.vendor_name or "",
            f"{item.po_amount:.2f}" if item.po_amount is not None else "",
            f"{item.invoice_amount:.2f}" if item.invoice_amount is not None else "",
            f"{item.amount_diff:.2f}" if item.amount_diff is not None else "",
            item.match_type or "",
            "是" if item.is_exception else "否",
            item.exception_type or "",
            item.status or "",
            item.rule_version or "",
            item.remarks or "",
            item.exception_remarks or "",
        ])

    return output.getvalue()


RELEASE_HEADER_FIELDS = {
    "放行包编号": "package_number",
    "放行包ID": "id",
    "批次ID": "batch_id",
    "状态": "status",
    "是否过期": "is_expired",
    "过期原因": "expire_reason",
    "批次状态快照": "batch_status",
    "应付合计": "payable_total",
    "匹配成功数": "matched_count",
    "异常数": "exception_count",
    "未匹配采购单数": "unmatched_po_count",
    "未匹配发票数": "unmatched_invoice_count",
    "容差百分比": "tolerance_pct",
    "容差绝对值": "tolerance_abs",
    "规则版本": "rule_version",
    "重算说明ID": "recalc_note_id",
    "重算说明版本": "recalc_note_version",
    "重算说明摘要": "recalc_note_summary",
    "巡检记录ID": "health_history_id",
    "巡检规则版本": "health_rule_version",
    "巡检阻断数": "health_blocker_count",
    "巡检警告数": "health_warning_count",
    "巡检提示数": "health_info_count",
    "导入方案ID": "import_plan_id",
    "经办人": "created_by",
    "提交人": "submitted_by",
    "审批人": "approved_by",
    "驳回人": "rejected_by",
    "驳回原因": "reject_reason",
    "撤销人": "revoked_by",
    "撤销原因": "revoke_reason",
    "内容哈希": "content_hash",
    "备注": "remarks",
}

RELEASE_DETAIL_FIELDS = [
    "序号", "匹配结果ID", "采购单号", "发票号", "供应商编码", "供应商名称",
    "采购金额", "发票金额", "金额差异", "匹配类型", "是否异常", "异常类型",
    "匹配状态", "规则版本", "匹配备注", "异常备注",
]


def import_release_csv(batch_id, csv_content, operator="system"):
    """回导放行包CSV，恢复快照。检查重复回导、跨批次回导、字段缺失"""
    batch = Batch.query.get(batch_id)
    if not batch:
        raise ValidationError(["批次不存在"])

    if isinstance(csv_content, bytes):
        csv_content = csv_content.decode("utf-8-sig")

    reader = csv.reader(io.StringIO(csv_content))
    rows = list(reader)

    in_header = False
    in_detail = False
    header_data = {}
    detail_items = []
    detail_headers = None

    for row in rows:
        if not row or not any(cell.strip() for cell in row):
            continue

        if row[0].strip() == "===== 放行包头段 =====":
            in_header = True
            in_detail = False
            continue
        if row[0].strip() == "===== 放行包明细 =====":
            in_header = False
            in_detail = True
            continue

        if in_header and len(row) >= 2:
            key = row[0].strip()
            value = row[1].strip() if len(row) > 1 else ""
            if key in RELEASE_HEADER_FIELDS:
                header_data[RELEASE_HEADER_FIELDS[key]] = value

        if in_detail:
            if detail_headers is None:
                detail_headers = [h.strip() for h in row]
                missing = [f for f in RELEASE_DETAIL_FIELDS if f not in detail_headers]
                if missing:
                    raise ValidationError([f"CSV 明细缺少必填列: {', '.join(missing)}"])
                continue

            item = {}
            for i, h in enumerate(detail_headers):
                item[h] = row[i].strip() if i < len(row) else ""
            detail_items.append(item)

    required_headers = ["package_number", "batch_id", "content_hash", "payable_total"]
    missing_headers = [k for k in required_headers if not header_data.get(k)]
    if missing_headers:
        field_map = {v: k for k, v in RELEASE_HEADER_FIELDS.items()}
        missing_names = [field_map[k] for k in missing_headers]
        raise ValidationError([f"CSV 头段缺少必填字段: {', '.join(missing_names)}"])

    csv_batch_id = None
    try:
        csv_batch_id = int(header_data["batch_id"])
    except (ValueError, TypeError):
        raise ValidationError([f"批次ID格式错误: {header_data['batch_id']}"])

    if csv_batch_id != batch_id:
        raise ValidationError([
            f"跨批次回导被拒绝：CSV 属于批次 #{csv_batch_id}，当前批次为 #{batch_id}"
        ])

    package_number = header_data["package_number"]
    existing = get_release_package_by_number(package_number)
    if existing:
        raise ValidationError([
            f"重复回导被拒绝：放行包 {package_number} 已存在 (#{existing.id})"
        ])

    if not detail_items:
        raise ValidationError(["CSV 没有明细数据"])

    check_release_expiry(batch_id, operator=operator)

    def safe_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def safe_int(val):
        if val is None or val == "":
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    pkg = ReleasePackage(
        batch_id=batch_id,
        package_number=package_number,
        status=RELEASE_STATUS_DRAFT,
        batch_status=header_data.get("batch_status"),
        payable_total=safe_float(header_data.get("payable_total")),
        matched_count=safe_int(header_data.get("matched_count")),
        exception_count=safe_int(header_data.get("exception_count")),
        unmatched_po_count=safe_int(header_data.get("unmatched_po_count")),
        unmatched_invoice_count=safe_int(header_data.get("unmatched_invoice_count")),
        tolerance_pct=safe_float(header_data.get("tolerance_pct")),
        tolerance_abs=safe_float(header_data.get("tolerance_abs")),
        rule_version=header_data.get("rule_version") or None,
        recalc_note_id=safe_int(header_data.get("recalc_note_id")),
        recalc_note_version=safe_int(header_data.get("recalc_note_version")),
        recalc_note_summary=header_data.get("recalc_note_summary") or None,
        health_history_id=safe_int(header_data.get("health_history_id")),
        health_rule_version=header_data.get("health_rule_version") or None,
        health_blocker_count=safe_int(header_data.get("health_blocker_count")) or 0,
        health_warning_count=safe_int(header_data.get("health_warning_count")) or 0,
        health_info_count=safe_int(header_data.get("health_info_count")) or 0,
        import_plan_id=safe_int(header_data.get("import_plan_id")),
        content_hash=header_data.get("content_hash") or None,
        is_expired=header_data.get("is_expired") == "是",
        expire_reason=header_data.get("expire_reason") or None,
        remarks=header_data.get("remarks") or None,
        created_by=header_data.get("created_by") or operator,
        submitted_by=header_data.get("submitted_by") or None,
        approved_by=header_data.get("approved_by") or None,
        rejected_by=header_data.get("rejected_by") or None,
        reject_reason=header_data.get("reject_reason") or None,
        revoked_by=header_data.get("revoked_by") or None,
        revoke_reason=header_data.get("revoke_reason") or None,
    )
    db.session.add(pkg)
    db.session.flush()

    for idx, row_data in enumerate(detail_items):
        item = ReleasePackageItem(
            release_package_id=pkg.id,
            match_result_id=safe_int(row_data.get("匹配结果ID")),
            po_number=row_data.get("采购单号") or None,
            invoice_number=row_data.get("发票号") or None,
            vendor_code=row_data.get("供应商编码") or None,
            vendor_name=row_data.get("供应商名称") or None,
            po_amount=safe_float(row_data.get("采购金额")),
            invoice_amount=safe_float(row_data.get("发票金额")),
            amount_diff=safe_float(row_data.get("金额差异")),
            match_type=row_data.get("匹配类型") or None,
            is_exception=row_data.get("是否异常") == "是",
            exception_type=row_data.get("异常类型") or None,
            status=row_data.get("匹配状态") or None,
            remarks=row_data.get("匹配备注") or None,
            exception_remarks=row_data.get("异常备注") or None,
            rule_version=row_data.get("规则版本") or None,
            item_order=idx,
        )
        db.session.add(item)

    log = AuditLog(
        batch_id=batch_id,
        action="RELEASE_IMPORT",
        detail=f"回导放行包 {package_number} (#{pkg.id})，共 {len(detail_items)} 条明细",
        operator=operator,
    )
    db.session.add(log)
    db.session.commit()
    db.session.refresh(pkg)

    return pkg


def list_release_audit_logs(batch_id=None, package_id=None, limit=50):
    query = AuditLog.query
    if batch_id is not None:
        query = query.filter_by(batch_id=batch_id)
    action_filter = "RELEASE_"
    query = query.filter(AuditLog.action.like(action_filter + "%"))
    if package_id is not None:
        query = query.filter(AuditLog.detail.like(f"%#{package_id}%"))
    return query.order_by(AuditLog.created_at.desc()).limit(limit).all()
